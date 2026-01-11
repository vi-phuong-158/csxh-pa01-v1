import unittest
import pandas as pd
import io
import openpyxl
from utils.bulk_import import export_error_excel

class TestSecurityUtils(unittest.TestCase):
    def test_excel_formula_injection_prevention(self):
        """Test that Excel Formula Injection is prevented in error reports."""
        # Create a dummy validation result with malicious input
        malicious_input = "=1+1"
        safe_input = "Normal Text"

        validation_results = {
            'doi_tuong': {
                'error_rows': [
                    {
                        'cccd': malicious_input,  # Injection point
                        'ho_ten': safe_input,
                        'LY_DO_LOI': 'Some error'
                    }
                ]
            }
        }

        # Generate Excel bytes
        excel_bytes = export_error_excel(validation_results)

        # Load the Excel file to check the cell value
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb['1. Đối tượng - LỖI']

        # Find column indices
        headers = [cell.value for cell in ws[1]]
        cccd_idx = headers.index('cccd') + 1
        hoten_idx = headers.index('ho_ten') + 1

        # Get values from row 2
        cccd_val = ws.cell(row=2, column=cccd_idx).value
        hoten_val = ws.cell(row=2, column=hoten_idx).value

        # Assertions
        # Malicious input should be escaped with single quote
        self.assertTrue(str(cccd_val).startswith("'="), f"Malicious input '{cccd_val}' was not escaped!")
        self.assertEqual(cccd_val, "'" + malicious_input)

        # Safe input should remain untouched (or at least not weirdly modified)
        self.assertEqual(hoten_val, safe_input)

if __name__ == '__main__':
    unittest.main()
