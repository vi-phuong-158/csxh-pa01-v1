# -*- coding: utf-8 -*-
"""
Script tạo dữ liệu giả định để test hệ thống Security Profile 360
LƯU Ý: Format Excel phải khớp với bulk_import.py:
- Row 1: Header (sẽ được dùng làm tên cột)  
- Row 2: Dòng mẫu/Sample (sẽ bị BỎ QUA bởi bulk_import)
- Row 3+: Dữ liệu thực tế
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import random

# Generate 10 sample persons


def generate_test_data():
    wb = Workbook()

    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea",
                              end_color="667eea", fill_type="solid")
    sample_font = Font(italic=True, color="888888")

    def style_header(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = max(
                15, len(header) + 3)

    def add_sample_row(ws, sample_data, row_num=2):
        """Add sample row (will be skipped by bulk_import)"""
        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = sample_font

    # ========== SHEET 1: ĐỐI TƯỢNG ==========
    ws1 = wb.active
    ws1.title = "1. Đối tượng"
    headers_1 = ["CCCD (*)", "Họ và tên (*)", "Ngày sinh (dd/mm/yyyy)",
                 "Giới tính", "Tỉnh/TP", "Xã/Phường",
                 "Phân loại nghề nghiệp", "Chi tiết nơi làm việc", "Ghi chú chung"]
    style_header(ws1, headers_1)

    # Sample row (row 2 - WILL BE SKIPPED by bulk_import)
    sample_1 = ["(Dòng mẫu - bị bỏ qua)", "(Họ tên mẫu)", "01/01/1990", "Nam", "Phú Thọ",
                "Phường Thanh Miếu", "Cơ quan nhà nước", "Chi tiết mẫu", "Ghi chú mẫu"]
    add_sample_row(ws1, sample_1)

    # Actual data (row 3+) - Using valid xã/phường from constants.py
    sample_persons = [
        ["012345678901", "Nguyễn Văn An", "15/03/1985", "Nam", "Phú Thọ", "Phường Thanh Miếu",
         "Cơ quan nhà nước", "Công an tỉnh Phú Thọ - Phòng PA01", "Cán bộ mẫu 1"],
        ["012345678902", "Trần Thị Bình", "22/07/1990", "Nữ", "Phú Thọ", "Phường Vân Phú",
         "Tự do", "Kinh doanh online", ""],
        ["012345678903", "Lê Hoàng Cường", "08/11/1978", "Nam", "Phú Thọ", "Xã Sơn Lương",
         "Doanh nghiệp tư nhân", "Giám đốc Công ty TNHH Minh Cường", "Có quan hệ quốc tế"],
        ["012345678904", "Phạm Thị Dung", "30/01/1995", "Nữ", "Phú Thọ", "Phường Phong Châu",
         "FDI", "Kỹ sư Samsung Bắc Ninh", "Du học Hàn Quốc 2018-2020"],
        ["012345678905", "Hoàng Minh Đức", "05/09/1982", "Nam", "Phú Thọ", "Xã Thanh Sơn",
         "Tự do", "Thợ xây dựng", "Từng lao động Đài Loan"],
        ["012345678906", "Vũ Thị Hoa", "12/05/1988", "Nữ", "Phú Thọ", "Phường Nông Trang",
         "Cơ quan nhà nước", "Giáo viên THPT Chuyên Hùng Vương", ""],
        ["012345678907", "Đặng Văn Giang", "25/12/1975", "Nam", "Phú Thọ", "Xã Phú Mỹ",
         "Nông nghiệp", "Nông dân", "Kết hôn với người Trung Quốc"],
        ["012345678908", "Bùi Thị Hạnh", "18/08/1992", "Nữ", "Phú Thọ", "Phường Việt Trì",
         "NGO", "Nhân viên World Vision", ""],
        ["012345678909", "Ngô Quốc Hùng", "03/04/1980", "Nam", "Phú Thọ", "Xã Hương Cần",
         "Doanh nghiệp tư nhân", "Chủ nhà hàng Hải Sản", ""],
        ["012345678910", "Lý Thị Kim", "28/06/1998", "Nữ", "Phú Thọ", "Phường Âu Cơ",
         "Tự do", "Content Creator", "Du học Nhật Bản 2020-2024"],
    ]

    for row in sample_persons:
        ws1.append(row)

    # ========== SHEET 2: LIÊN HỆ ==========
    ws2 = wb.create_sheet("2. Liên hệ")
    headers_2 = ["CCCD (*)", "Loại liên hệ (*)", "Giá trị (*)", "Ghi chú"]
    style_header(ws2, headers_2)

    # Sample row (will be skipped)
    add_sample_row(ws2, ["(CCCD mẫu)", "SĐT", "0123456789", "Ghi chú mẫu"])

    sample_contacts = [
        ["012345678901", "SĐT", "0987654321", "Di động cá nhân"],
        ["012345678901", "Email", "nguyenvanan@gmail.com", "Email chính"],
        ["012345678902", "SĐT", "0912345678", ""],
        ["012345678902", "Zalo", "0912345678", ""],
        ["012345678903", "SĐT", "0909123456", "Số công ty"],
        ["012345678903", "Email", "cuong.minhcuong@company.vn", ""],
        ["012345678904", "SĐT", "0978111222", ""],
        ["012345678904", "Facebook", "facebook.com/dungpham95", ""],
        ["012345678905", "SĐT", "0867123456", ""],
        ["012345678906", "SĐT", "0356789012", ""],
        ["012345678906", "Email", "hoa.vuthi@edu.vn", "Email công việc"],
        ["012345678907", "SĐT", "0823456789", ""],
        ["012345678908", "SĐT", "0945678901", ""],
        ["012345678908", "Email", "hanh.bui@worldvision.org", ""],
        ["012345678909", "SĐT", "0789012345", "Số nhà hàng"],
        ["012345678910", "SĐT", "0976543210", ""],
        ["012345678910", "Instagram", "@lykim.creator", "Kênh chính"],
    ]

    for row in sample_contacts:
        ws2.append(row)

    # ========== SHEET 3: TÀI CHÍNH ==========
    ws3 = wb.create_sheet("3. Tài chính")
    headers_3 = ["CCCD (*)", "Ngân hàng (*)",
                 "Số tài khoản (*)", "Chủ tài khoản", "Ghi chú"]
    style_header(ws3, headers_3)

    # Sample row (will be skipped)
    add_sample_row(ws3, ["(CCCD mẫu)", "Vietcombank",
                   "0123456789", "CHU TK", "Ghi chú mẫu"])

    sample_banking = [
        ["012345678901", "Vietcombank", "1234567890123", "NGUYEN VAN AN", "Lương"],
        ["012345678902", "BIDV", "31410001234567", "TRAN THI BINH", ""],
        ["012345678903", "Techcombank", "19032123456789",
            "LE HOANG CUONG", "TK công ty"],
        ["012345678903", "VietinBank", "104881234567",
            "LE HOANG CUONG", "TK cá nhân"],
        ["012345678904", "MB Bank", "0123456789012", "PHAM THI DUNG", ""],
        ["012345678905", "Agribank", "4500201234567", "HOANG MINH DUC", ""],
        ["012345678906", "TPBank", "02345678901", "VU THI HOA", ""],
        ["012345678907", "Vietcombank", "0071234567890", "DANG VAN GIANG", ""],
        ["012345678908", "ACB", "123456789", "BUI THI HANH", ""],
        ["012345678909", "Sacombank", "060012345678",
            "NGO QUOC HUNG", "TK nhà hàng"],
        ["012345678910", "VPBank", "12345678901234",
            "LY THI KIM", "Nhận tiền YouTube"],
    ]

    for row in sample_banking:
        ws3.append(row)

    # ========== SHEET 4: PHƯƠNG TIỆN ==========
    ws4 = wb.create_sheet("4. Phương tiện")
    headers_4 = ["CCCD (*)", "Loại xe (*)",
                 "Biển kiểm soát (*)", "Tên xe", "Ghi chú"]
    style_header(ws4, headers_4)

    # Sample row (will be skipped)
    add_sample_row(ws4, ["(CCCD mẫu)", "Xe máy",
                   "19A-12345", "Honda Vision", "Ghi chú mẫu"])

    sample_vehicles = [
        ["012345678901", "Ô tô con", "19A-12345", "Toyota Vios", "Xe cá nhân"],
        ["012345678902", "Xe máy", "19B1-56789", "Honda Vision", ""],
        ["012345678903", "Ô tô con", "30A-99999", "Mercedes C300", "Xe công ty"],
        ["012345678903", "Ô tô tải", "19C-88888",
            "Hyundai Porter", "Xe vận chuyển"],
        ["012345678904", "Xe máy", "19C2-11111", "Yamaha Exciter", ""],
        ["012345678905", "Xe máy", "19D1-33333", "Honda Wave", ""],
        ["012345678906", "Ô tô con", "19A-55555", "Mazda 3", ""],
        ["012345678907", "Xe máy", "19E1-44444", "Honda Dream", ""],
        ["012345678908", "Xe máy", "19B1-22222", "Honda SH", ""],
        ["012345678909", "Ô tô con", "19A-66666", "Ford Ranger", ""],
        ["012345678909", "Xe máy", "19C1-77777", "Honda Air Blade", ""],
        ["012345678910", "Xe máy", "19B2-12121", "Vespa Sprint", ""],
    ]

    for row in sample_vehicles:
        ws4.append(row)

    # ========== SHEET 5: HỒ SƠ ĐẶC THÙ ==========
    ws5 = wb.create_sheet("5. Hồ sơ CSXH")
    headers_5 = ["CCCD (*)", "Loại hình CSXH (*)", "Tên đối tác/Tổ chức (*)",
                 "Quốc gia (*)", "Số hộ chiếu", "Tình trạng", "Năm bắt đầu", "Năm kết thúc", "Ghi chú"]
    style_header(ws5, headers_5)

    # Sample row (will be skipped)
    add_sample_row(ws5, ["(CCCD mẫu)", "Hon_Nhan_NN",
                   "Tên mẫu", "Việt Nam", "", "", "", "", ""])

    sample_csxh = [
        ["012345678903", "Hon_Nhan_NN", "Maria Rodriguez", "Tây Ban Nha", "ES123456789",
         "Kết hôn hợp pháp", "2015", "", "Quen khi du lịch"],
        ["012345678904", "Hoc_Tap_Cong_Tac_NN", "Đại học Seoul", "Hàn Quốc", "",
         "Đã hoàn thành", "2018", "2020", "Học bổng chính phủ"],
        ["012345678004", "Lam_Viec_NN", "Samsung Electronics", "Hàn Quốc", "",
         "Đang làm việc", "2020", "", "Kỹ sư phần mềm"],
        ["012345678905", "Hoc_Tap_Cong_Tac_NN", "Công ty XYZ", "Đài Loan", "",
         "Đã về nước", "2010", "2015", "Xuất khẩu lao động"],
        ["012345678907", "Hon_Nhan_NN", "Zhang Wei", "Trung Quốc", "CN987654321",
         "Kết hôn hợp pháp", "2018", "", ""],
        ["012345678908", "Lam_Viec_NN", "World Vision Vietnam", "Mỹ", "",
         "Đang làm việc", "2019", "", "NGO"],
        ["012345678910", "Hoc_Tap_Cong_Tac_NN", "Đại học Tokyo", "Nhật Bản", "",
         "Đã hoàn thành", "2020", "2024", "Học bổng MEXT"],
    ]

    for row in sample_csxh:
        ws5.append(row)

    # Save file
    output_path = "test_data_v2.xlsx"
    wb.save(output_path)
    print(f"✅ Đã tạo file: {output_path}")
    print(f"   - 10 đối tượng mẫu")
    print(f"   - {len(sample_contacts)} liên hệ")
    print(f"   - {len(sample_banking)} tài khoản ngân hàng")
    print(f"   - {len(sample_vehicles)} phương tiện")
    print(f"   - {len(sample_csxh)} hồ sơ đặc thù CSXH")
    return output_path


if __name__ == "__main__":
    generate_test_data()
