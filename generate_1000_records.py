# -*- coding: utf-8 -*-
"""
Script tạo 1000 hồ sơ giả định cho hệ thống Security Profile 360
Dựa trên cấu trúc bulk import đã định nghĩa
"""

import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# ============================================
# DANH SÁCH DỮ LIỆU MẪU
# ============================================

# Họ phổ biến Việt Nam
HO_COMMON = [
    "Nguyễn", "Trần", "Lê", "Phạm", "Hoàng", "Huỳnh", "Phan", "Vũ", "Võ",
    "Đặng", "Bùi", "Đỗ", "Hồ", "Ngô", "Dương", "Lý", "Đinh", "Lương",
    "Trương", "Đào", "Cao", "Tạ", "Lưu", "Tô", "Hà", "Chu"
]

# Tên đệm phổ biến
TEN_DEM = [
    "Văn", "Thị", "Hữu", "Đức", "Ngọc", "Minh", "Hoàng", "Thanh", "Quốc",
    "Xuân", "Hồng", "Kim", "Phương", "Thành", "Mạnh", "Tuấn", "Anh",
    "Bích", "Diệu", "Thu", "Hà", "Mai", "Lan", "Linh", "Hương", "Tú"
]

# Tên phổ biến Nam
TEN_NAM = [
    "Anh", "Bình", "Cường", "Dũng", "Đức", "Giang", "Hải", "Hiếu", "Hoàng",
    "Hùng", "Khải", "Khoa", "Kiên", "Long", "Minh", "Nam", "Nghĩa", "Nhân",
    "Phong", "Phú", "Quang", "Quyết", "Sơn", "Thắng", "Thành", "Thiện",
    "Trung", "Tuấn", "Tùng", "Việt", "Vũ", "Lâm", "Huy", "Bảo", "Toàn"
]

# Tên phổ biến Nữ
TEN_NU = [
    "Anh", "Bình", "Chi", "Diệp", "Dung", "Giang", "Hà", "Hạnh", "Hằng",
    "Hiền", "Hoa", "Hương", "Lan", "Linh", "Loan", "Mai", "Ngọc", "Nhung",
    "Phương", "Quỳnh", "Thảo", "Thu", "Thủy", "Trang", "Trinh", "Uyên",
    "Vân", "Yến", "Hồng", "Vy", "Nhi", "My", "Trâm", "Như", "Thúy"
]

# Danh sách 15 phường
DANH_SACH_PHUONG = [
    "Phường Âu Cơ", "Phường Hòa Bình", "Phường Kỳ Sơn", "Phường Nông Trang",
    "Phường Phong Châu", "Phường Phú Thọ", "Phường Phúc Yên", "Phường Tân Hòa",
    "Phường Thanh Miếu", "Phường Thống Nhất", "Phường Vân Phú", "Phường Việt Trì",
    "Phường Vĩnh Phúc", "Phường Vĩnh Yên", "Phường Xuân Hòa"
]

# Danh sách một số xã
DANH_SACH_XA = [
    "Xã An Bình", "Xã Bản Nguyên", "Xã Bình Phú", "Xã Cao Dương", "Xã Cẩm Khê",
    "Xã Chân Mộng", "Xã Dân Chủ", "Xã Đại Đình", "Xã Đào Xá", "Xã Hạ Hòa",
    "Xã Hiền Lương", "Xã Hoàng Cương", "Xã Hùng Việt", "Xã Hy Cương", "Xã Lâm Thao",
    "Xã Liên Hòa", "Xã Long Cốc", "Xã Minh Đài", "Xã Ngọc Sơn", "Xã Phú Khê",
    "Xã Thanh Ba", "Xã Thanh Sơn", "Xã Thu Cúc", "Xã Tiên Lương", "Xã Văn Lang",
    "Xã Vĩnh Chân", "Xã Xuân Đài", "Xã Yên Lạc", "Xã Yên Sơn", "Xã Tu Vũ"
]

DANH_SACH_XA_PHU_THO = DANH_SACH_PHUONG + DANH_SACH_XA

# Phân loại nghề nghiệp
PHAN_LOAI_NGHE_NGHIEP = [
    "Cơ quan nhà nước", "Lao động tự do", "Doanh nghiệp tư nhân", "Nông nghiệp",
    "FDI", "NGO", "Học sinh/Sinh viên", "Hưu trí", "Thất nghiệp", "Khác"
]

# Chi tiết nơi làm việc theo phân loại
CHI_TIET_NGHE_NGHIEP = {
    "Cơ quan nhà nước": [
        "Công an tỉnh Phú Thọ", "UBND tỉnh Phú Thọ", "Sở Tài chính", "Sở Giáo dục",
        "UBND thành phố Việt Trì", "Chi cục Thuế", "Kho bạc Nhà nước", "Văn phòng HĐND",
        "Sở Công thương", "Sở Nông nghiệp", "Bộ Công an", "UBND huyện Thanh Ba"
    ],
    "Lao động tự do": [
        "Kinh doanh buôn bán nhỏ", "Xe ôm công nghệ", "Thợ xây dựng", "Thợ điện nước",
        "Bán hàng chợ", "Shipper", "Thợ may", "Sửa chữa điện tử", "Buôn bán online"
    ],
    "Doanh nghiệp tư nhân": [
        "Công ty TNHH ABC", "Công ty CP XYZ", "Doanh nghiệp tư nhân Minh Anh",
        "Công ty TNHH MTV Hùng Cường", "Công ty CP Thực phẩm Việt", "Ngân hàng TMCP"
    ],
    "Nông nghiệp": [
        "Trồng chè", "Chăn nuôi lợn", "Trồng lúa", "Nuôi trồng thủy sản",
        "Trồng rau sạch", "Chăn nuôi gia cầm", "Làm vườn", "Trồng cây ăn quả"
    ],
    "FDI": [
        "Samsung Electronics Vietnam", "LG Display", "Canon Vietnam", "Panasonic",
        "Honda Vietnam", "Toyota", "Hyundai", "Intel Products Vietnam"
    ],
    "NGO": [
        "World Vision", "Care International", "Save the Children", "WWF Vietnam",
        "Oxfam", "Plan International", "UNDP Vietnam"
    ],
    "Học sinh/Sinh viên": [
        "Trường THPT Chuyên Hùng Vương", "Đại học Hùng Vương", "Cao đẳng Y tế Phú Thọ",
        "Trường THCS Thanh Miếu", "Đại học Bách Khoa Hà Nội", "ĐH Kinh tế Quốc dân"
    ],
    "Hưu trí": ["Hưu trí tại nhà", "Chăm cháu", "Nghỉ hưu tại quê"],
    "Thất nghiệp": ["Đang tìm việc", "Nội trợ"],
    "Khác": ["Tự do", "Không xác định"]
}

# Loại liên hệ
LOAI_LIEN_HE = ["SĐT", "Email", "Facebook", "Zalo", "Telegram", "Instagram"]

# Danh sách ngân hàng
NGAN_HANG = [
    "Vietcombank", "Vietinbank", "BIDV", "Agribank", "Techcombank",
    "MB Bank", "ACB", "Sacombank", "VPBank", "TPBank"
]

# Loại xe
LOAI_XE = ["Ô tô", "Xe máy", "Ô tô con", "Ô tô tải", "Xe đạp điện"]

# Quan hệ thân nhân
QUAN_HE = ["Bố đẻ", "Mẹ đẻ", "Vợ", "Chồng", "Con trai",
           "Con gái", "Anh trai", "Em gái", "Ông nội", "Bà nội"]

# Quốc gia
QUOC_GIA = ["Trung Quốc", "Hàn Quốc", "Nhật Bản", "Đài Loan",
            "Thái Lan", "Malaysia", "Singapore", "Mỹ", "Úc", "Đức"]

# ============================================
# HÀM TẠO DỮ LIỆU NGẪU NHIÊN
# ============================================


def generate_cccd():
    """Tạo CCCD 12 số ngẫu nhiên (bắt đầu với 025 - Phú Thọ)"""
    # Tạo 9 số ngẫu nhiên và đảm bảo luôn có đủ 9 chữ số
    random_part = str(random.randint(0, 999999999)).zfill(9)
    return f"025{random_part}"


def generate_name(gioi_tinh):
    """Tạo họ tên ngẫu nhiên theo giới tính"""
    ho = random.choice(HO_COMMON)
    dem = random.choice(TEN_DEM)
    ten = random.choice(TEN_NAM if gioi_tinh == "Nam" else TEN_NU)
    return f"{ho} {dem} {ten}"


def generate_birth_date(min_age=18, max_age=70):
    """Tạo ngày sinh ngẫu nhiên"""
    today = datetime.now()
    days_old = random.randint(min_age * 365, max_age * 365)
    birth_date = today - timedelta(days=days_old)
    return birth_date.strftime("%d/%m/%Y")


def generate_phone():
    """Tạo số điện thoại ngẫu nhiên"""
    prefixes = ["090", "091", "093", "094", "096", "097", "098", "099",
                "032", "033", "034", "035", "036", "037", "038", "039", "086"]
    return f"{random.choice(prefixes)}{random.randint(1000000, 9999999)}"


def generate_email(name):
    """Tạo email từ tên"""
    domains = ["gmail.com", "yahoo.com", "outlook.com", "fpt.vn"]
    name_part = name.lower().replace(" ", ".")
    # Remove Vietnamese diacritics (simplified)
    import unicodedata
    name_part = ''.join(c for c in unicodedata.normalize(
        'NFD', name_part) if unicodedata.category(c) != 'Mn')
    name_part = name_part.replace("đ", "d")
    return f"{name_part}{random.randint(1, 999)}@{random.choice(domains)}"


def generate_bank_account():
    """Tạo số tài khoản ngân hàng ngẫu nhiên"""
    return str(random.randint(1000000000, 9999999999999))


def generate_license_plate():
    """Tạo biển số xe ngẫu nhiên (Phú Thọ: 19)"""
    letters = "ABCDEFGHKLMNPRSTUVXYZ"
    return f"19{random.choice(letters)}-{random.randint(10000, 99999)}"


def generate_vehicle_name(loai_xe):
    """Tạo tên phương tiện"""
    cars = ["Toyota Vios", "Honda City", "Hyundai Accent",
            "Kia Morning", "Mazda 3", "VinFast Lux"]
    bikes = ["Honda Wave", "Honda SH", "Yamaha Exciter",
             "Suzuki Raider", "Honda Winner", "VinFast Klara"]
    if "Ô tô" in loai_xe:
        return random.choice(cars)
    return random.choice(bikes)

# ============================================
# TẠO DỮ LIỆU CHO TỪNG SHEET
# ============================================


def generate_test_data(num_records=1000):
    """Tạo dữ liệu test cho 1000 hồ sơ"""

    print(f"📊 Bắt đầu tạo {num_records} hồ sơ giả định...")

    # Sheet 1: Đối tượng (bắt buộc)
    doi_tuong_data = []
    all_cccd = []

    for i in range(num_records):
        cccd = generate_cccd()
        # Đảm bảo CCCD không trùng
        while cccd in all_cccd:
            cccd = generate_cccd()
        all_cccd.append(cccd)

        gioi_tinh = random.choice(["Nam", "Nữ"])
        ho_ten = generate_name(gioi_tinh)
        ngay_sinh = generate_birth_date()
        xa_phuong = random.choice(DANH_SACH_XA_PHU_THO)
        phan_loai = random.choice(PHAN_LOAI_NGHE_NGHIEP)
        chi_tiet = random.choice(CHI_TIET_NGHE_NGHIEP.get(phan_loai, ["N/A"]))

        doi_tuong_data.append({
            "CCCD (*)": cccd,
            "Họ và tên (*)": ho_ten,
            "Ngày sinh (dd/mm/yyyy)": ngay_sinh,
            "Giới tính": gioi_tinh,
            "Tỉnh/TP": "Phú Thọ",
            "Xã/Phường": xa_phuong,
            "Phân loại nghề nghiệp": phan_loai,
            "Chi tiết nơi làm việc": chi_tiet,
            "Ghi chú chung": ""
        })

        if (i + 1) % 100 == 0:
            print(f"  ✅ Đã tạo {i + 1} đối tượng...")

    # Sheet 2: Liên hệ (1-3 liên hệ mỗi người, ~80% có liên hệ)
    lien_he_data = []
    for i, cccd in enumerate(all_cccd):
        if random.random() < 0.80:  # 80% có liên hệ
            # SĐT (hầu hết đều có)
            lien_he_data.append({
                "CCCD (*)": cccd,
                "Loại liên hệ": "SĐT",
                "Giá trị (*)": generate_phone(),
                "Ghi chú": "SĐT chính"
            })
            # Email (50% có)
            if random.random() < 0.50:
                ho_ten = doi_tuong_data[i]["Họ và tên (*)"]
                lien_he_data.append({
                    "CCCD (*)": cccd,
                    "Loại liên hệ": "Email",
                    "Giá trị (*)": generate_email(ho_ten),
                    "Ghi chú": ""
                })
            # Zalo/Facebook (30% có)
            if random.random() < 0.30:
                lien_he_data.append({
                    "CCCD (*)": cccd,
                    "Loại liên hệ": random.choice(["Zalo", "Facebook"]),
                    "Giá trị (*)": generate_phone(),
                    "Ghi chú": ""
                })

    print(f"  ✅ Đã tạo {len(lien_he_data)} liên hệ...")

    # Sheet 3: Thân nhân (1-3 thân nhân mỗi người, ~70% có thân nhân)
    than_nhan_data = []
    for i, cccd in enumerate(all_cccd):
        if random.random() < 0.70:  # 70% có thân nhân
            num_relatives = random.randint(1, 3)
            for _ in range(num_relatives):
                gioi_tinh_tn = random.choice(["Nam", "Nữ"])
                than_nhan_data.append({
                    "CCCD (*)": cccd,
                    "Họ tên thân nhân": generate_name(gioi_tinh_tn),
                    "Quan hệ": random.choice(QUAN_HE),
                    "Năm sinh": str(random.randint(1940, 2010)),
                    "Nghề nghiệp/Nơi làm việc": random.choice(["Hưu trí", "Nội trợ", "Kinh doanh", "CBCC", "Nông dân"]),
                    "Địa chỉ": random.choice(DANH_SACH_XA_PHU_THO) + ", Phú Thọ",
                    "Ghi chú": ""
                })

    print(f"  ✅ Đã tạo {len(than_nhan_data)} thân nhân...")

    # Sheet 4: Tài chính (~40% có tài khoản ngân hàng)
    tai_chinh_data = []
    for i, cccd in enumerate(all_cccd):
        if random.random() < 0.40:  # 40% có tài khoản
            ho_ten = doi_tuong_data[i]["Họ và tên (*)"].upper()
            # Loại bỏ dấu
            import unicodedata
            ho_ten = ''.join(c for c in unicodedata.normalize(
                'NFD', ho_ten) if unicodedata.category(c) != 'Mn')
            ho_ten = ho_ten.replace("Đ", "D")

            tai_chinh_data.append({
                "CCCD (*)": cccd,
                "Ngân hàng": random.choice(NGAN_HANG),
                "Số tài khoản (*)": generate_bank_account(),
                "Chủ tài khoản": ho_ten,
                "Ghi chú": ""
            })

    print(f"  ✅ Đã tạo {len(tai_chinh_data)} tài khoản ngân hàng...")

    # Sheet 5: Phương tiện (~30% có phương tiện đăng ký)
    phuong_tien_data = []
    for cccd in all_cccd:
        if random.random() < 0.30:  # 30% có phương tiện
            loai_xe = random.choice(LOAI_XE)
            phuong_tien_data.append({
                "CCCD (*)": cccd,
                "Loại xe": loai_xe,
                "Biển kiểm soát (*)": generate_license_plate(),
                "Tên phương tiện": generate_vehicle_name(loai_xe),
                "Ghi chú": ""
            })

    print(f"  ✅ Đã tạo {len(phuong_tien_data)} phương tiện...")

    # Sheet 6: Quá trình hoạt động (~50% có thông tin)
    qua_trinh_data = []
    for i, cccd in enumerate(all_cccd):
        if random.random() < 0.50:  # 50% có quá trình
            num_entries = random.randint(1, 3)
            birth_year = int(
                doi_tuong_data[i]["Ngày sinh (dd/mm/yyyy)"].split("/")[2])

            for j in range(num_entries):
                start_year = birth_year + 18 + j * 5
                end_year = start_year + random.randint(3, 5)
                if end_year > 2024:
                    end_year = "nay"

                activities = [
                    "Học sinh THPT", "Sinh viên Đại học", "Nhân viên văn phòng",
                    "Kỹ sư phần mềm", "Giáo viên", "Bác sĩ", "Công nhân",
                    "Cán bộ xã/phường", "Kinh doanh tự do", "Nghĩa vụ quân sự"
                ]

                qua_trinh_data.append({
                    "CCCD (*)": cccd,
                    "Thời gian (từ năm-đến năm)": f"{start_year}-{end_year}",
                    "Nội dung hoạt động": random.choice(activities),
                    "Ghi chú": ""
                })

    print(f"  ✅ Đã tạo {len(qua_trinh_data)} quá trình hoạt động...")

    # Sheet 6: Hồ sơ đặc thù CSXH (~15% có yếu tố nước ngoài)
    # Format: CCCD, Loại hình, Quốc tịch, Tên tổ chức, Từ năm, Đến năm, Nội dung, Cơ quan XM, Kết quả, Ghi chú
    ho_so_dac_thu_data = []
    valid_loai_hinh = ['Hon_Nhan_NN', 'Lam_Viec_NN',
                       'Hoc_Tap_Cong_Tac_NN', 'Vi_Pham_NN', 'Xac_Minh']

    for cccd in all_cccd:
        if random.random() < 0.15:  # 15% có yếu tố đặc thù
            loai_hinh = random.choice(valid_loai_hinh)
            quoc_gia = random.choice(QUOC_GIA)
            nam_bat_dau = str(random.randint(2010, 2023))
            nam_ket_thuc = str(random.randint(
                int(nam_bat_dau), 2024)) if random.random() > 0.3 else ""

            if loai_hinh == "Hon_Nhan_NN":
                ten_nguoi = random.choice(
                    ["WANG Xiaoming", "KIM Min-jun", "TANAKA Yuki", "CHEN Wei", "LEE Jae-hyun"])
                ho_so_dac_thu_data.append({
                    "CCCD (*)": cccd,
                    "Loại hình (*)": loai_hinh,
                    "Quốc tịch/Quốc gia": quoc_gia,
                    "Tên tổ chức/Người nước ngoài": ten_nguoi,
                    "Thời gian (từ năm)": nam_bat_dau,
                    "Thời gian (đến năm)": nam_ket_thuc,
                    "Nội dung chi tiết": f"Kết hôn với công dân {quoc_gia}",
                    "Cơ quan xác minh": "",
                    "Kết quả": "",
                    "Ghi chú": ""
                })
            elif loai_hinh == "Lam_Viec_NN":
                companies = ["Samsung Electronics Vietnam", "LG Display", "Canon Vietnam",
                             "Panasonic VN", "Intel Products Vietnam", "Honda Vietnam"]
                ho_so_dac_thu_data.append({
                    "CCCD (*)": cccd,
                    "Loại hình (*)": loai_hinh,
                    "Quốc tịch/Quốc gia": quoc_gia,
                    "Tên tổ chức/Người nước ngoài": random.choice(companies),
                    "Thời gian (từ năm)": nam_bat_dau,
                    "Thời gian (đến năm)": nam_ket_thuc,
                    "Nội dung chi tiết": "Làm việc tại doanh nghiệp FDI",
                    "Cơ quan xác minh": "",
                    "Kết quả": "",
                    "Ghi chú": ""
                })
            elif loai_hinh == "Hoc_Tap_Cong_Tac_NN":
                truong = ["Đại học Tokyo", "Đại học Seoul",
                          "Đại học Bắc Kinh", "MIT", "Đại học Melbourne"]
                ho_so_dac_thu_data.append({
                    "CCCD (*)": cccd,
                    "Loại hình (*)": loai_hinh,
                    "Quốc tịch/Quốc gia": quoc_gia,
                    "Tên tổ chức/Người nước ngoài": random.choice(truong),
                    "Thời gian (từ năm)": nam_bat_dau,
                    "Thời gian (đến năm)": nam_ket_thuc,
                    "Nội dung chi tiết": f"Du học tại {quoc_gia}",
                    "Cơ quan xác minh": "",
                    "Kết quả": "",
                    "Ghi chú": ""
                })
            elif loai_hinh == "Vi_Pham_NN":
                vi_pham = ["Cư trú bất hợp pháp", "Vi phạm lao động",
                           "Vi phạm giao thông", "Ở quá hạn visa"]
                ho_so_dac_thu_data.append({
                    "CCCD (*)": cccd,
                    "Loại hình (*)": loai_hinh,
                    "Quốc tịch/Quốc gia": quoc_gia,
                    "Tên tổ chức/Người nước ngoài": "",
                    "Thời gian (từ năm)": nam_bat_dau,
                    "Thời gian (đến năm)": nam_ket_thuc,
                    "Nội dung chi tiết": random.choice(vi_pham),
                    "Cơ quan xác minh": "",
                    "Kết quả": "Đã xử lý",
                    "Ghi chú": "Đã về nước"
                })
            else:  # Xac_Minh
                co_quan = ["PA01", "Công an tỉnh",
                           "Sở Công thương", "Bộ Công an"]
                ket_qua = random.choice(
                    ["Đủ điều kiện", "Không đủ điều kiện", "Đang xác minh"])
                ho_so_dac_thu_data.append({
                    "CCCD (*)": cccd,
                    "Loại hình (*)": loai_hinh,
                    "Quốc tịch/Quốc gia": "",
                    "Tên tổ chức/Người nước ngoài": "",
                    "Thời gian (từ năm)": nam_bat_dau,
                    "Thời gian (đến năm)": nam_ket_thuc,
                    "Nội dung chi tiết": "Xác minh bổ nhiệm cán bộ",
                    "Cơ quan xác minh": random.choice(co_quan),
                    "Kết quả": ket_qua,
                    "Ghi chú": ""
                })

    print(f"  ✅ Đã tạo {len(ho_so_dac_thu_data)} hồ sơ đặc thù...")

    return {
        "doi_tuong": pd.DataFrame(doi_tuong_data),
        "lien_he": pd.DataFrame(lien_he_data),
        "than_nhan": pd.DataFrame(than_nhan_data),
        "tai_chinh": pd.DataFrame(tai_chinh_data),
        "phuong_tien": pd.DataFrame(phuong_tien_data),
        "qua_trinh_hoat_dong": pd.DataFrame(qua_trinh_data),
        "ho_so_dac_thu": pd.DataFrame(ho_so_dac_thu_data)
    }


def save_to_excel(data_dict, filename="test_data_1000.xlsx"):
    """Lưu dữ liệu vào file Excel multi-sheet theo format bulk import"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # Sheet config matching bulk_import format
    sheet_config = {
        "doi_tuong": {
            "name": "1. Đối tượng",
            "headers": ["CCCD (*)", "Họ và tên (*)", "Ngày sinh (dd/mm/yyyy)",
                        "Giới tính", "Tỉnh/TP", "Xã/Phường",
                        "Phân loại nghề nghiệp", "Chi tiết nơi làm việc", "Ghi chú chung"],
            "sample": ["001234567890", "Nguyễn Văn A", "01/01/1990", "Nam",
                       "Phú Thọ", "Phường Thanh Miếu", "Cơ quan nhà nước",
                       "Công an tỉnh Phú Thọ", "Ghi chú mẫu"]
        },
        "lien_he": {
            "name": "2. Liên hệ",
            "headers": ["CCCD (*)", "Loại liên hệ", "Giá trị (*)", "Ghi chú"],
            "sample": ["001234567890", "SĐT", "0912345678", "SĐT chính"]
        },
        "than_nhan": {
            "name": "3. Thân nhân",
            "headers": ["CCCD (*)", "Họ tên thân nhân", "Quan hệ", "Năm sinh",
                        "Nghề nghiệp/Nơi làm việc", "Địa chỉ", "Ghi chú"],
            "sample": ["001234567890", "Nguyễn Văn B", "Bố đẻ", "1960",
                       "Hưu trí", "Việt Trì, Phú Thọ", ""]
        },
        "tai_chinh": {
            "name": "4. Tài chính",
            "headers": ["CCCD (*)", "Ngân hàng", "Số tài khoản (*)", "Chủ tài khoản", "Ghi chú"],
            "sample": ["001234567890", "Vietcombank", "1234567890123", "NGUYEN VAN A", "TK chính"]
        },
        "phuong_tien": {
            "name": "5. Phương tiện",
            "headers": ["CCCD (*)", "Loại xe", "Biển kiểm soát (*)", "Tên phương tiện", "Ghi chú"],
            "sample": ["001234567890", "Ô tô", "19A-12345", "Toyota Vios 2022", "Xe cá nhân"]
        },
        "ho_so_dac_thu": {
            "name": "6. Hồ sơ CSXH",
            "headers": ["CCCD (*)", "Loại hình (*)", "Quốc tịch/Quốc gia", "Tên tổ chức/Người nước ngoài",
                        "Thời gian (từ năm)", "Thời gian (đến năm)", "Nội dung chi tiết",
                        "Cơ quan xác minh", "Kết quả", "Ghi chú"],
            "sample": ["001234567890", "Hon_Nhan_NN", "Trung Quốc", "WANG Xiaoming",
                       "2020", "", "Kết hôn với công dân TQ", "", "", ""]
        }
    }

    first_sheet = True
    for key in ["doi_tuong", "lien_he", "than_nhan", "tai_chinh", "phuong_tien", "ho_so_dac_thu"]:
        if key not in data_dict or data_dict[key].empty:
            continue

        config = sheet_config.get(key)
        if not config:
            continue

        if first_sheet:
            ws = wb.active
            ws.title = config["name"]
            first_sheet = False
        else:
            ws = wb.create_sheet(title=config["name"])

        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="667eea", end_color="667eea", fill_type="solid")

        # Write headers
        for col, header in enumerate(config["headers"], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = max(
                15, len(header) + 5)

        # Write sample row (row 2 - sẽ bị bulk import bỏ qua)
        for col, value in enumerate(config["sample"], 1):
            ws.cell(row=2, column=col, value=value)

        # Write actual data starting from row 3
        df = data_dict[key]
        for row_idx, row_data in enumerate(df.values, 3):
            for col_idx, value in enumerate(row_data, 1):
                # Đảm bảo CCCD là text với leading zeros
                if col_idx == 1:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = str(value)
                    cell.number_format = '@'  # Text format
                else:
                    ws.cell(row=row_idx, column=col_idx,
                            value=value if pd.notna(value) else "")

    wb.save(filename)
    print(f"\n✅ Đã lưu file: {filename}")
    print(f"📊 Thống kê:")
    for key, df in data_dict.items():
        if key in sheet_config:
            print(f"   - {key}: {len(df)} bản ghi")


# ============================================
# MAIN
# ============================================
if __name__ == "__main__":
    print("=" * 50)
    print("🎲 SCRIPT TẠO DỮ LIỆU THỬ NGHIỆM")
    print("   Hệ thống Security Profile PA01")
    print("=" * 50)

    # Tạo 1000 hồ sơ
    data = generate_test_data(1000)

    # Lưu vào file Excel
    output_file = "test_data_1000.xlsx"
    save_to_excel(data, output_file)

    print("\n" + "=" * 50)
    print("🎉 HOÀN TẤT!")
    print(f"📁 File output: {output_file}")
    print("💡 Sử dụng file này để import vào hệ thống qua chức năng 'Nhập Excel'")
    print("=" * 50)
