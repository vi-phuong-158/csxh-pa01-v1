# -*- coding: utf-8 -*-
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def export_error_excel(validation_results):
    """
    Tạo file Excel chứa các dòng lỗi với cột lý do để user fix
    Returns: bytes của file Excel hoặc None nếu không có lỗi
    """
    wb = Workbook()

    # Style cho header
    header_font = Font(bold=True, color="FFFFFF")
    error_fill = PatternFill(start_color="dc3545",
                             end_color="dc3545", fill_type="solid")
    header_fill = PatternFill(start_color="667eea",
                              end_color="667eea", fill_type="solid")

    has_errors = False
    first_sheet = True

    # Sheet tên và dữ liệu
    sheet_configs = [
        ('1. Đối tượng - LỖI', 'doi_tuong'),
        ('2. Liên hệ - LỖI', 'lien_he'),
        ('3. Tài chính - LỖI', 'tai_chinh'),
        ('4. Phương tiện - LỖI', 'phuong_tien'),
        ('5. CSXH - LỖI', 'ho_so_dac_thu'),
        ('6. Quá trình - LỖI', 'qua_trinh_hoat_dong'), # Added this
        ('7. Thân nhân - LỖI', 'than_nhan'), # Added this
    ]

    for sheet_name, key in sheet_configs:
        error_rows = validation_results.get(key, {}).get('error_rows', [])
        if error_rows:
            has_errors = True

            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = wb.create_sheet(sheet_name)

            # Tạo DataFrame từ error_rows
            df = pd.DataFrame(error_rows)

            # Đảm bảo cột LY_DO_LOI ở cuối
            if 'LY_DO_LOI' in df.columns:
                cols = [c for c in df.columns if c !=
                        'LY_DO_LOI'] + ['LY_DO_LOI']
                df = df[cols]

            # Viết header
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                if col_name == 'LY_DO_LOI':
                    cell.fill = error_fill
                else:
                    cell.fill = header_fill
                cell.font = header_font
                ws.column_dimensions[cell.column_letter].width = max(
                    15, len(str(col_name)) + 5)

            # Viết dữ liệu - sử dụng enumerate để có row number chính xác
            for row_num, (_, row) in enumerate(df.iterrows(), start=2):
                for col_idx, value in enumerate(row.values, 1):
                    # Security: Sanitize potential formula injection
                    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')):
                        value = "'" + value

                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    # Highlight cột lý do lỗi
                    if df.columns[col_idx - 1] == 'LY_DO_LOI':
                        cell.font = Font(color="dc3545", bold=True)

    if not has_errors:
        return None

    # Lưu vào buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()
