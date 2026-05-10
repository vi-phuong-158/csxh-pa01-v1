# -*- coding: utf-8 -*-
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .constants import CSXH_TEMPLATES, TEMPLATE_DEFINITIONS

def style_header_row(ws, headers):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea",
                              end_color="667eea", fill_type="solid")
    thin_border = Border(left=Side('thin'), right=Side(
        'thin'), top=Side('thin'), bottom=Side('thin'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(
            15, len(header) + 5)


def create_excel_template(import_type="all", csxh_type=None):
    """
    Tạo file Excel mẫu
    Args:
        import_type: 'all' (5 sheet) hoặc key trong TEMPLATE_DEFINITIONS ('doi_tuong', 'lien_he'...)
        csxh_type: Loại CSXH cụ thể (nếu import_type='ho_so_dac_thu')
    """
    wb = Workbook()

    if import_type == "all":
        create_full_template(wb, csxh_type)
    elif import_type == "ho_so_dac_thu":
        # Tạo sheet CSXH lẻ
        if csxh_type and csxh_type in CSXH_TEMPLATES:
            tpl = CSXH_TEMPLATES[csxh_type]
            ws = wb.active
            ws.title = "Hồ sơ CSXH"
            style_header_row(ws, tpl['headers'])
            ws.append(tpl['sample'])
            ws.cell(row=4, column=1, value=f"Loại hình: {csxh_type}")
        else:
            # CSXH Tổng hợp
            create_csxh_general_sheet(wb)
    elif import_type in TEMPLATE_DEFINITIONS:
        # Tạo sheet đơn cho các loại khác
        tpl = TEMPLATE_DEFINITIONS[import_type]
        ws = wb.active
        ws.title = tpl['name']
        style_header_row(ws, tpl['headers'])
        ws.append(tpl['sample'])
        ws.cell(row=5, column=1,
                value="(*) cột bắt buộc. CCCD dùng để định danh đối tượng.")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def create_full_template(wb, csxh_type):
    # ========== SHEET 1: ĐỐI TƯỢNG ==========
    ws1 = wb.active
    ws1.title = "1. Đối tượng"
    # Dòng mẫu
    ws1.append(["001234567890", "Nguyễn Văn A", "01/01/1990", "Nam",
                "Phú Thọ", "Phường Thanh Miếu", "Cơ quan nhà nước",
                "Công an tỉnh Phú Thọ", "Ghi chú mẫu"])
    ws1.append(["001234567891", "Trần Thị B", "15/05/1985", "Nữ",
                "Phú Thọ", "Phường Gia Cẩm", "Tự do",
                "Buôn bán tự do", ""])

    # Ghi chú
    ws1.cell(row=5, column=1,
             value="Lưu ý: (*) là trường bắt buộc. CCCD phải đủ 12 số.")
    
    # Apply style manually for sheet 1 since it's hardcoded here
    headers_1 = ["CCCD (*)", "Họ và tên (*)", "Ngày sinh (dd/mm/yyyy)",
                 "Giới tính", "Tỉnh/TP", "Xã/Phường",
                 "Phân loại nghề nghiệp", "Chi tiết nơi làm việc", "Ghi chú chung"]
    style_header_row(ws1, headers_1)


    # ========== SHEET 2: LIÊN HỆ ==========
    ws2 = wb.create_sheet("2. Liên hệ")
    headers_2 = [
        "CCCD (*)", "Loại liên hệ", "Giá trị (*)", "Ghi chú"
    ]
    style_header_row(ws2, headers_2)
    ws2.append(["001234567890", "Số điện thoại", "0912345678", "SĐT chính"])
    ws2.append(["001234567890", "Facebook", "facebook.com/nguyenvana", ""])
    ws2.append(["001234567891", "Số điện thoại", "0987654321", ""])
    ws2.cell(row=6, column=1,
             value="Loại liên hệ: Số điện thoại, Email, Facebook, Zalo, Telegram, Khác")

    # ========== SHEET 3: TÀI CHÍNH ==========
    ws3 = wb.create_sheet("3. Tài chính")
    headers_3 = [
        "CCCD (*)", "Ngân hàng", "Số tài khoản (*)", "Chủ tài khoản", "Ghi chú"
    ]
    style_header_row(ws3, headers_3)
    ws3.append(["001234567890", "Vietcombank",
               "1234567890123", "NGUYEN VAN A", "TK chính"])
    ws3.append(["001234567890", "Techcombank",
               "9876543210", "NGUYEN VAN A", "TK phụ"])

    # ========== SHEET 4: PHƯƠNG TIỆN ==========
    ws4 = wb.create_sheet("4. Phương tiện")
    headers_4 = [
        "CCCD (*)", "Loại xe", "Biển kiểm soát (*)", "Tên phương tiện", "Ghi chú"
    ]
    style_header_row(ws4, headers_4)
    ws4.append(["001234567890", "Ô tô", "19A-12345",
               "Toyota Vios 2022", "Xe cá nhân"])
    ws4.append(["001234567891", "Xe máy", "19B1-67890", "Honda Wave", ""])
    ws4.cell(row=5, column=1, value="Loại xe: Ô tô, Xe máy, Xe tải, Xe khách, Khác")

    # ========== SHEET 6: QUAN HỆ ==========
    ws6 = wb.create_sheet("6. Quan hệ")
    headers_6 = TEMPLATE_DEFINITIONS['than_nhan']['headers']
    style_header_row(ws6, headers_6)
    ws6.append(["001234567890", "001234567891", "Nguyễn Văn B", "Cha-Con",
                "1960", "Hưu trí", "Việt Trì, Phú Thọ", ""])
    ws6.append(["001234567890", "", "John Smith", "Đồng nghiệp",
                "1985", "Kỹ sư", "", "Không có CCCD → lưu dạng ghi chú"])
    ws6.cell(row=5, column=1,
             value=TEMPLATE_DEFINITIONS['than_nhan'].get('note', ''))

    # ========== SHEET 7: QUÁ TRÌNH HOẠT ĐỘNG ==========
    ws7 = wb.create_sheet("7. Quá trình hoạt động")
    headers_7 = TEMPLATE_DEFINITIONS['qua_trinh_hoat_dong']['headers']
    style_header_row(ws7, headers_7)
    ws7.append(["001234567890", "2010-2015", "Học sinh trường THPT Chuyên Hùng Vương", ""])

    # ========== SHEET 5: HỒ SƠ CSXH (Luôn để cuối hoặc sau cùng) ==========
    if csxh_type and csxh_type in CSXH_TEMPLATES:
        # Tạo sheet theo loại cụ thể
        template = CSXH_TEMPLATES[csxh_type]
        ws5 = wb.create_sheet(f"5. {template['name']}")
        style_header_row(ws5, template['headers'])
        ws5.append(template['sample'])
        ws5.cell(row=4, column=1, value=f"Loại hình: {csxh_type}")
    else:
        create_csxh_general_sheet(wb)

def create_csxh_general_sheet(wb):
    ws5 = wb.create_sheet("5. Hồ sơ CSXH (Tổng hợp)")
    headers_5 = [
        "CCCD (*)", "Loại hình (*)",
        "Quốc tịch/Quốc gia", "Tên tổ chức/Người nước ngoài",
        "Thời gian (từ năm)", "Thời gian (đến năm)",
        "Nội dung chi tiết", "Cơ quan xác minh", "Kết quả", "Ghi chú"
    ]
    style_header_row(ws5, headers_5)

    # Mẫu cho từng loại
    ws5.append(["001234567890", "Hon_Nhan_NN", "Trung Quốc", "WANG Xiaoming",
                "2020", "", "Kết hôn với công dân TQ", "", "", ""])

    ws5.cell(row=4, column=1, value="--- HƯỚNG DẪN LOẠI HÌNH ---")
    ws5.cell(row=5, column=1,
             value="Hon_Nhan_NN | Lam_Viec_NN | Hoc_Tap_Cong_Tac_NN | Vi_Pham_NN | Xac_Minh")
