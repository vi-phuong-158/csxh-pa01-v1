# File: backend/routes/bao_cao.py
import logging
from datetime import datetime, timedelta
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import and_, func, select
from sqlalchemy.orm import Session

from backend.constants import PHAN_LOAI_NGHE_NGHIEP, LOAI_HINH_DAC_THU
from backend.db.session import get_db
from backend.deps import require_login
from backend.models.models import (
    DoiTuong, LienHe, TaiChinh, HoSoDacThu,
    NhanThan, PhuongTien, QuaTrinhHoatDong,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/bao-cao", tags=["bao-cao"])
templates = Jinja2Templates(directory="frontend/templates")


# ── Private helpers ───────────────────────────────────────────────────────────

def _parse_dates(
    tu_ngay: Optional[str], den_ngay: Optional[str]
) -> tuple[Optional[datetime], Optional[datetime]]:
    dt_from = datetime.strptime(tu_ngay, "%Y-%m-%d") if tu_ngay else None
    dt_to = (
        datetime.strptime(den_ngay, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        if den_ngay
        else None
    )
    return dt_from, dt_to


def _query_stats(
    db: Session,
    dt_from: Optional[datetime],
    dt_to: Optional[datetime],
    phan_loai: Optional[str],
    loai_hs_dac_thu: Optional[str] = None,
) -> dict:
    def base_filter():
        conds = [DoiTuong.is_draft == False]
        if dt_from:
            conds.append(DoiTuong.created_at >= dt_from)
        if dt_to:
            conds.append(DoiTuong.created_at <= dt_to)
        if phan_loai:
            conds.append(DoiTuong.phan_loai_nghe_nghiep == phan_loai)
        if loai_hs_dac_thu:
            if loai_hs_dac_thu == "ALL":
                conds.append(DoiTuong.cccd.in_(select(HoSoDacThu.cccd).scalar_subquery()))
            else:
                conds.append(DoiTuong.cccd.in_(select(HoSoDacThu.cccd).where(HoSoDacThu.loai_hinh == loai_hs_dac_thu).scalar_subquery()))
        return and_(*conds)

    total = db.execute(
        select(func.count(DoiTuong.cccd)).where(base_filter())
    ).scalar_one()

    draft_count = db.execute(
        select(func.count(DoiTuong.cccd)).where(DoiTuong.is_draft == True)
    ).scalar_one()

    co_sdt = db.execute(
        select(func.count(DoiTuong.cccd)).where(
            base_filter(),
            DoiTuong.cccd.in_(select(LienHe.cccd).distinct()),
        )
    ).scalar_one()

    co_stk = db.execute(
        select(func.count(DoiTuong.cccd)).where(
            base_filter(),
            DoiTuong.cccd.in_(select(TaiChinh.cccd).distinct()),
        )
    ).scalar_one()

    phan_loai_rows = db.execute(
        select(DoiTuong.phan_loai_nghe_nghiep, func.count(DoiTuong.cccd))
        .where(base_filter())
        .group_by(DoiTuong.phan_loai_nghe_nghiep)
        .order_by(func.count(DoiTuong.cccd).desc())
    ).all()
    by_phan_loai = {(r[0] or "Không rõ"): r[1] for r in phan_loai_rows}

    month_col = func.strftime("%Y-%m", DoiTuong.created_at)
    month_rows = db.execute(
        select(month_col.label("month"), func.count(DoiTuong.cccd).label("cnt"))
        .where(base_filter(), DoiTuong.created_at.isnot(None))
        .group_by(month_col)
        .order_by(month_col)
        .limit(24)
    ).all()
    by_month = [{"month": r[0], "count": r[1]} for r in month_rows if r[0]]

    gioi_tinh_rows = db.execute(
        select(DoiTuong.gioi_tinh, func.count(DoiTuong.cccd))
        .where(base_filter())
        .group_by(DoiTuong.gioi_tinh)
    ).all()
    by_gioi_tinh = {(r[0] or "Không rõ"): r[1] for r in gioi_tinh_rows}

    dia_ban_rows = db.execute(
        select(DoiTuong.dia_chi_xa, func.count(DoiTuong.cccd))
        .where(
            base_filter(),
            DoiTuong.dia_chi_xa.isnot(None),
            DoiTuong.dia_chi_xa != "",
        )
        .group_by(DoiTuong.dia_chi_xa)
        .order_by(func.count(DoiTuong.cccd).desc())
        .limit(10)
    ).all()
    by_dia_ban = {r[0]: r[1] for r in dia_ban_rows}

    table = [
        {
            "phan_loai": loai,
            "count": cnt,
            "pct": round(cnt / total * 100, 1) if total > 0 else 0,
        }
        for loai, cnt in by_phan_loai.items()
    ]

    return {
        "summary": {
            "total": total,
            "draft": draft_count,
            "co_sdt": co_sdt,
            "co_stk": co_stk,
        },
        "by_phan_loai": by_phan_loai,
        "by_month": by_month,
        "by_gioi_tinh": by_gioi_tinh,
        "by_dia_ban": by_dia_ban,
        "table": table,
    }


def _build_xlsx(stats: dict, filter_lines: list, generated_at: datetime, detailed_records: list = None) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Shared styles ─────────────────────────────────────────────────────────
    _thin = Side(style="thin", color="E2E8F0")
    BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center", indent=1)
    RIGHT  = Alignment(horizontal="right",  vertical="center")

    TITLE_FONT = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    TITLE_FILL = PatternFill(fill_type="solid", fgColor="1E40AF")
    TITLE_ALIGN = Alignment(horizontal="center", vertical="center")

    SEC_FONT  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    SEC_FILL  = PatternFill(fill_type="solid", fgColor="0F766E")
    SEC_ALIGN = Alignment(horizontal="left", vertical="center", indent=1)

    HDR_FONT  = Font(bold=True, color="1E40AF", size=10, name="Calibri")
    HDR_FILL  = PatternFill(fill_type="solid", fgColor="DBEAFE")

    LABEL_FONT = Font(bold=True, color="334155", size=10, name="Calibri")
    LABEL_FILL = PatternFill(fill_type="solid", fgColor="F1F5F9")
    DATA_FONT  = Font(color="1E293B", size=10, name="Calibri")
    ALT_FILL   = PatternFill(fill_type="solid", fgColor="F8FAFC")
    NUM_FONT   = Font(bold=True, color="1E40AF", size=11, name="Calibri")
    LINK_FONT  = Font(color="1D4ED8", size=10, name="Calibri", underline="single")

    # ── Helper utilities ──────────────────────────────────────────────────────
    def _c(ws, row, col, value=None, *, font=None, fill=None, align=None,
           border=None, fmt=None):
        cell = ws.cell(row=row, column=col, value=value)
        if font   is not None: cell.font      = font
        if fill   is not None: cell.fill      = fill
        if align  is not None: cell.alignment = align
        if border is not None: cell.border    = border
        if fmt    is not None: cell.number_format = fmt
        return cell

    def _title_row(ws, row, ncols, text, is_section=False):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        _c(ws, row, 1, text,
           font=SEC_FONT if is_section else TITLE_FONT,
           fill=SEC_FILL if is_section else TITLE_FILL,
           align=SEC_ALIGN if is_section else TITLE_ALIGN)
        ws.row_dimensions[row].height = 30 if not is_section else 22

    def _col_headers(ws, row, headers, widths=None):
        for i, h in enumerate(headers, 1):
            _c(ws, row, i, h, font=HDR_FONT, fill=HDR_FILL,
               align=Alignment(horizontal="center", vertical="center", wrap_text=True),
               border=BORDER)
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[row].height = 22

    def _data_row(ws, row, values, aligns=None):
        fill = ALT_FILL if row % 2 == 0 else None
        for i, v in enumerate(values, 1):
            a = (aligns[i - 1] if aligns else None) or LEFT
            _c(ws, row, i, v, font=DATA_FONT, fill=fill, align=a, border=BORDER)

    # Pull data
    summary    = stats["summary"]
    by_phan_loai = stats["by_phan_loai"]
    by_month   = stats["by_month"]
    by_gioi_tinh = stats["by_gioi_tinh"]
    by_dia_ban  = stats["by_dia_ban"]
    table      = stats["table"]
    total      = summary["total"]

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Tóm tắt
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Tóm tắt"
    ws1.sheet_view.showGridLines = False

    _title_row(ws1, 1, 3, "BÁO CÁO THỐNG KÊ ĐỐI TƯỢNG")

    ws1.merge_cells("A2:C2")
    _c(ws1, 2, 1, "Security Profile 360 · PA01",
       font=Font(italic=True, color="64748B", size=10, name="Calibri"),
       fill=PatternFill(fill_type="solid", fgColor="F8FAFC"),
       align=CENTER)
    ws1.row_dimensions[2].height = 16

    # Metadata rows
    meta = [
        ("Ngày xuất báo cáo", generated_at.strftime("%d/%m/%Y %H:%M")),
        ("Bộ lọc áp dụng",   " | ".join(filter_lines)),
    ]
    for i, (lbl, val) in enumerate(meta, 3):
        _c(ws1, i, 1, lbl, font=LABEL_FONT, fill=LABEL_FILL, align=LEFT, border=BORDER)
        ws1.merge_cells(f"B{i}:C{i}")
        _c(ws1, i, 2, val, font=DATA_FONT, fill=ALT_FILL, align=LEFT, border=BORDER)

    # Summary section
    _title_row(ws1, 6, 3, "  SỐ LIỆU TỔNG HỢP", is_section=True)

    kpi_rows = [
        ("Tổng hồ sơ hoàn chỉnh",  summary["total"]),
        ("Hồ sơ đang soạn thảo",   summary["draft"]),
        ("Có số điện thoại",        summary["co_sdt"]),
        ("Có tài khoản ngân hàng",  summary["co_stk"]),
    ]
    for i, (lbl, val) in enumerate(kpi_rows, 7):
        _c(ws1, i, 1, lbl, font=LABEL_FONT, fill=LABEL_FILL, align=LEFT, border=BORDER)
        ws1.merge_cells(f"B{i}:C{i}")
        _c(ws1, i, 2, val, font=NUM_FONT,   fill=None, align=CENTER, border=BORDER)

    # Giới tính section
    _title_row(ws1, 12, 3, "  PHÂN BỐ GIỚI TÍNH", is_section=True)
    _col_headers(ws1, 13, ["Giới tính", "Số hồ sơ", "Tỷ lệ (%)"])
    for i, (gt, cnt) in enumerate(by_gioi_tinh.items(), 1):
        pct = round(cnt / total * 100, 1) if total else 0
        _data_row(ws1, 13 + i, [gt, cnt, pct], aligns=[LEFT, CENTER, CENTER])
        ws1.cell(row=13 + i, column=3).number_format = '0.0"%"'

    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 14

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Phân loại nghề nghiệp
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Phân loại NNg")
    ws2.sheet_view.showGridLines = False

    _title_row(ws2, 1, 4, "PHÂN BỐ THEO PHÂN LOẠI NGHỀ NGHIỆP")
    _col_headers(ws2, 2,
                 ["STT", "Phân loại nghề nghiệp", "Số hồ sơ", "Tỷ lệ (%)"],
                 widths=[6, 36, 14, 14])

    for i, row in enumerate(table, 1):
        r = i + 2
        fill = ALT_FILL if i % 2 == 0 else None
        _c(ws2, r, 1, i,             font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)
        _c(ws2, r, 2, row["phan_loai"], font=DATA_FONT, fill=fill, align=LEFT,   border=BORDER)
        _c(ws2, r, 3, row["count"],     font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)
        _c(ws2, r, 4, row["pct"],       font=DATA_FONT, fill=fill, align=CENTER, border=BORDER,
           fmt='0.0"%"')

    if table:
        ws2.freeze_panes = "A3"
        ws2.auto_filter.ref = f"A2:D{2 + len(table)}"
        ws2.conditional_formatting.add(
            f"D3:D{2 + len(table)}",
            DataBarRule(start_type="num", start_value=0,
                        end_type="num",   end_value=100,
                        color="3B82F6"),
        )

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Biến động theo tháng
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Biến động tháng")
    ws3.sheet_view.showGridLines = False

    _title_row(ws3, 1, 3, "BIẾN ĐỘNG NHẬP HỒ SƠ THEO THÁNG")
    _col_headers(ws3, 2, ["STT", "Tháng", "Số hồ sơ nhập mới"],
                 widths=[6, 18, 22])

    for i, m in enumerate(by_month, 1):
        r = i + 2
        fill = ALT_FILL if i % 2 == 0 else None
        _c(ws3, r, 1, i,          font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)
        _c(ws3, r, 2, m["month"], font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)
        _c(ws3, r, 3, m["count"], font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)

    if by_month:
        ws3.freeze_panes = "A3"
        ws3.auto_filter.ref = f"A2:C{2 + len(by_month)}"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Địa bàn
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Địa bàn")
    ws4.sheet_view.showGridLines = False

    _title_row(ws4, 1, 3, "TOP 10 ĐỊA BÀN CÓ NHIỀU HỒ SƠ NHẤT")
    _col_headers(ws4, 2, ["STT", "Địa bàn (Xã/Phường)", "Số hồ sơ"],
                 widths=[6, 36, 16])

    for i, (db_name, cnt) in enumerate(by_dia_ban.items(), 1):
        r = i + 2
        fill = ALT_FILL if i % 2 == 0 else None
        _c(ws4, r, 1, i,       font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)
        _c(ws4, r, 2, db_name, font=DATA_FONT, fill=fill, align=LEFT,   border=BORDER)
        _c(ws4, r, 3, cnt,     font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)

    if by_dia_ban:
        ws4.freeze_panes = "A3"
        ws4.auto_filter.ref = f"A2:C{2 + len(by_dia_ban)}"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — Thông tin cơ bản (mở rộng từ 11 → 17 cột)
    # ═══════════════════════════════════════════════════════════════════════════
    if detailed_records is not None:
        ws5 = wb.create_sheet("Thông tin cơ bản")
        ws5.sheet_view.showGridLines = False

        _title_row(ws5, 1, 17, "DANH SÁCH CHI TIẾT ĐỐI TƯỢNG — THÔNG TIN CƠ BẢN")

        headers5 = [
            "STT", "CCCD", "Họ tên", "Ngày sinh", "Giới tính",
            "Địa chỉ xã", "Địa chỉ tỉnh", "Phân loại NNg",
            "Chi tiết nghề nghiệp", "Số điện thoại", "Email",
            "Liên hệ khác", "Tài khoản ngân hàng", "Phương tiện",
            "Hồ sơ đặc thù", "Ghi chú", "Ngày tạo hồ sơ",
        ]
        _col_headers(ws5, 2, headers5, widths=[
            6, 16, 25, 12, 10, 22, 16, 18,
            25, 18, 22,
            28, 28, 22,
            25, 25, 14,
        ])

        for i, dt in enumerate(detailed_records, 1):
            r = i + 2
            fill = ALT_FILL if i % 2 == 0 else None

            ngay_sinh_str = dt.ngay_sinh.strftime("%d/%m/%Y") if dt.ngay_sinh else ""
            created_str = dt.created_at.strftime("%d/%m/%Y") if dt.created_at else ""

            sdts = [lh.gia_tri for lh in dt.lien_he if lh.loai_lien_he == "SĐT" and lh.gia_tri]
            emails = [lh.gia_tri for lh in dt.lien_he if lh.loai_lien_he == "Email" and lh.gia_tri]
            others = [
                f"{lh.loai_lien_he}: {lh.gia_tri}"
                for lh in dt.lien_he
                if lh.loai_lien_he not in ("SĐT", "Email") and lh.gia_tri
            ]
            tk_list = [
                f"{tc.ngan_hang or ''} - {tc.so_tai_khoan or ''}".strip(" -")
                for tc in dt.tai_chinh if tc.so_tai_khoan
            ]
            xe_list = [
                f"{pt.loai_xe or ''} - {pt.bien_kiem_soat or ''}".strip(" -")
                for pt in dt.phuong_tien if pt.bien_kiem_soat or pt.loai_xe
            ]
            hs_list = [
                LOAI_HINH_DAC_THU.get(hs.loai_hinh, hs.loai_hinh)
                for hs in dt.ho_so_dac_thu
            ]

            vals5 = [
                i,
                dt.cccd,
                dt.ho_ten or "",
                ngay_sinh_str,
                dt.gioi_tinh or "",
                dt.dia_chi_xa or "",
                dt.dia_chi_tinh or "",
                dt.phan_loai_nghe_nghiep or "",
                dt.chi_tiet_nghe_nghiep or "",
                "; ".join(sdts),
                "; ".join(emails),
                "; ".join(others),
                "; ".join(tk_list),
                "; ".join(xe_list),
                "; ".join(hs_list),
                dt.ghi_chu_chung or "",
                created_str,
            ]
            aligns5 = [
                CENTER, CENTER, LEFT, CENTER, CENTER, LEFT, LEFT, LEFT,
                LEFT, CENTER, LEFT,
                LEFT, LEFT, LEFT,
                LEFT, LEFT, CENTER,
            ]
            for col_idx, v in enumerate(vals5, 1):
                _c(ws5, r, col_idx, v, font=DATA_FONT, fill=fill,
                   align=aligns5[col_idx - 1], border=BORDER)

        if detailed_records:
            ws5.freeze_panes = "A3"
            ws5.auto_filter.ref = f"A2:Q{2 + len(detailed_records)}"

    # Mapping cccd → row trong Sheet 5 (dùng để tạo hyperlink từ sheet phụ)
    cccd_to_row5: dict = {
        dt.cccd: idx + 2
        for idx, dt in enumerate(detailed_records or [], 1)
    }

    def _cccd_link(ws, row, cccd, ho_ten):
        """Ghi ô CCCD có hyperlink về Sheet 5, và ô Họ tên kế bên."""
        fill = ALT_FILL if row % 2 == 0 else None
        target_row = cccd_to_row5.get(cccd)
        cell = ws.cell(row=row, column=2, value=cccd)
        cell.font   = LINK_FONT
        cell.fill   = fill or PatternFill()
        cell.alignment = CENTER
        cell.border = BORDER
        if target_row:
            cell.hyperlink = f"#'Thông tin cơ bản'!B{target_row}"
        _c(ws, row, 3, ho_ten, font=DATA_FONT, fill=fill, align=LEFT, border=BORDER)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 6 — Nhân thân
    # ═══════════════════════════════════════════════════════════════════════════
    if detailed_records:
        nhan_than_rows = [
            (dt, nt) for dt in detailed_records for nt in dt.nhan_than
        ]
        if nhan_than_rows:
            ws6 = wb.create_sheet("Nhân thân")
            ws6.sheet_view.showGridLines = False
            _title_row(ws6, 1, 12, "DANH SÁCH NHÂN THÂN")
            _col_headers(ws6, 2, [
                "STT", "CCCD chủ hồ sơ", "Họ tên chủ hồ sơ",
                "Quan hệ", "Họ tên nhân thân", "CCCD nhân thân",
                "Ngày sinh", "Giới tính", "Nghề nghiệp",
                "Nơi ở", "Địa chỉ tỉnh", "Ghi chú",
            ], widths=[6, 16, 25, 14, 25, 16, 12, 10, 20, 28, 16, 25])
            for i, (dt, nt) in enumerate(nhan_than_rows, 1):
                r = i + 2
                fill = ALT_FILL if i % 2 == 0 else None
                ns = nt.ngay_sinh.strftime("%d/%m/%Y") if nt.ngay_sinh else ""
                _c(ws6, r, 1, i, font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)
                _cccd_link(ws6, r, dt.cccd, dt.ho_ten or "")
                rest = [
                    nt.loai_quan_he or "", nt.ho_ten or "", nt.cccd_nhan_than or "",
                    ns, nt.gioi_tinh or "", nt.nghe_nghiep or "",
                    nt.noi_o or "", nt.dia_chi_tinh or "", nt.ghi_chu or "",
                ]
                al_rest = [LEFT, LEFT, CENTER, CENTER, CENTER, LEFT, LEFT, LEFT, LEFT]
                for ci, v in enumerate(rest, 4):
                    _c(ws6, r, ci, v, font=DATA_FONT, fill=fill, align=al_rest[ci - 4], border=BORDER)
            ws6.freeze_panes = "A3"
            ws6.auto_filter.ref = f"A2:L{2 + len(nhan_than_rows)}"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 7 — Phương tiện
    # ═══════════════════════════════════════════════════════════════════════════
    if detailed_records:
        pt_rows = [(dt, pt) for dt in detailed_records for pt in dt.phuong_tien]
        if pt_rows:
            ws7 = wb.create_sheet("Phương tiện")
            ws7.sheet_view.showGridLines = False
            _title_row(ws7, 1, 7, "DANH SÁCH PHƯƠNG TIỆN")
            _col_headers(ws7, 2, [
                "STT", "CCCD chủ hồ sơ", "Họ tên chủ hồ sơ",
                "Loại xe", "Biển kiểm soát", "Tên phương tiện", "Ghi chú",
            ], widths=[6, 16, 25, 16, 16, 25, 25])
            for i, (dt, pt) in enumerate(pt_rows, 1):
                r = i + 2
                fill = ALT_FILL if i % 2 == 0 else None
                _c(ws7, r, 1, i, font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)
                _cccd_link(ws7, r, dt.cccd, dt.ho_ten or "")
                rest = [pt.loai_xe or "", pt.bien_kiem_soat or "", pt.ten_phuong_tien or "", pt.ghi_chu or ""]
                al_rest = [LEFT, CENTER, LEFT, LEFT]
                for ci, v in enumerate(rest, 4):
                    _c(ws7, r, ci, v, font=DATA_FONT, fill=fill, align=al_rest[ci - 4], border=BORDER)
            ws7.freeze_panes = "A3"
            ws7.auto_filter.ref = f"A2:G{2 + len(pt_rows)}"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 8 — Tài khoản ngân hàng
    # ═══════════════════════════════════════════════════════════════════════════
    if detailed_records:
        tc_rows = [(dt, tc) for dt in detailed_records for tc in dt.tai_chinh]
        if tc_rows:
            ws8 = wb.create_sheet("Tài khoản ngân hàng")
            ws8.sheet_view.showGridLines = False
            _title_row(ws8, 1, 7, "DANH SÁCH TÀI KHOẢN NGÂN HÀNG")
            _col_headers(ws8, 2, [
                "STT", "CCCD chủ hồ sơ", "Họ tên chủ hồ sơ",
                "Ngân hàng", "Số tài khoản", "Chủ tài khoản", "Ghi chú",
            ], widths=[6, 16, 25, 22, 20, 25, 25])
            for i, (dt, tc) in enumerate(tc_rows, 1):
                r = i + 2
                fill = ALT_FILL if i % 2 == 0 else None
                _c(ws8, r, 1, i, font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)
                _cccd_link(ws8, r, dt.cccd, dt.ho_ten or "")
                rest = [tc.ngan_hang or "", tc.so_tai_khoan or "", tc.chu_tai_khoan or "", tc.ghi_chu or ""]
                al_rest = [LEFT, CENTER, LEFT, LEFT]
                for ci, v in enumerate(rest, 4):
                    _c(ws8, r, ci, v, font=DATA_FONT, fill=fill, align=al_rest[ci - 4], border=BORDER)
            ws8.freeze_panes = "A3"
            ws8.auto_filter.ref = f"A2:G{2 + len(tc_rows)}"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 9 — Quá trình hoạt động
    # ═══════════════════════════════════════════════════════════════════════════
    if detailed_records:
        qt_rows = [(dt, qt) for dt in detailed_records for qt in dt.qua_trinh]
        if qt_rows:
            ws9 = wb.create_sheet("Quá trình hoạt động")
            ws9.sheet_view.showGridLines = False
            _title_row(ws9, 1, 8, "QUÁ TRÌNH HOẠT ĐỘNG")
            _col_headers(ws9, 2, [
                "STT", "CCCD chủ hồ sơ", "Họ tên chủ hồ sơ",
                "Thời gian", "Ngày bắt đầu", "Ngày kết thúc",
                "Nội dung", "Ghi chú",
            ], widths=[6, 16, 25, 16, 14, 14, 40, 25])
            for i, (dt, qt) in enumerate(qt_rows, 1):
                r = i + 2
                fill = ALT_FILL if i % 2 == 0 else None
                bat_dau = qt.ngay_bat_dau.strftime("%d/%m/%Y") if qt.ngay_bat_dau else ""
                ket_thuc = qt.ngay_ket_thuc.strftime("%d/%m/%Y") if qt.ngay_ket_thuc else ""
                _c(ws9, r, 1, i, font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)
                _cccd_link(ws9, r, dt.cccd, dt.ho_ten or "")
                rest = [qt.thoi_gian or "", bat_dau, ket_thuc, qt.noi_dung or "", qt.ghi_chu or ""]
                al_rest = [LEFT, CENTER, CENTER, LEFT, LEFT]
                for ci, v in enumerate(rest, 4):
                    _c(ws9, r, ci, v, font=DATA_FONT, fill=fill, align=al_rest[ci - 4], border=BORDER)
            ws9.freeze_panes = "A3"
            ws9.auto_filter.ref = f"A2:H{2 + len(qt_rows)}"

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 10 — Hồ sơ đặc thù
    # ═══════════════════════════════════════════════════════════════════════════
    if detailed_records:
        hs_rows = [(dt, hs) for dt in detailed_records for hs in dt.ho_so_dac_thu]
        if hs_rows:
            ws10 = wb.create_sheet("Hồ sơ đặc thù")
            ws10.sheet_view.showGridLines = False
            _title_row(ws10, 1, 6, "HỒ SƠ ĐẶC THÙ")
            _col_headers(ws10, 2, [
                "STT", "CCCD chủ hồ sơ", "Họ tên chủ hồ sơ",
                "Loại hình", "Nội dung chi tiết", "Ghi chú",
            ], widths=[6, 16, 25, 28, 40, 25])
            for i, (dt, hs) in enumerate(hs_rows, 1):
                r = i + 2
                fill = ALT_FILL if i % 2 == 0 else None
                loai_str = LOAI_HINH_DAC_THU.get(hs.loai_hinh, hs.loai_hinh or "")
                _c(ws10, r, 1, i, font=DATA_FONT, fill=fill, align=CENTER, border=BORDER)
                _cccd_link(ws10, r, dt.cccd, dt.ho_ten or "")
                rest = [loai_str, hs.noi_dung_chi_tiet or "", hs.ghi_chu or ""]
                al_rest = [LEFT, LEFT, LEFT]
                for ci, v in enumerate(rest, 4):
                    _c(ws10, r, ci, v, font=DATA_FONT, fill=fill, align=al_rest[ci - 4], border=BORDER)
            ws10.freeze_panes = "A3"
            ws10.auto_filter.ref = f"A2:F{2 + len(hs_rows)}"

    # ── Serialize ─────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("", response_class=HTMLResponse)
def bao_cao_page(request: Request, user: dict = Depends(require_login)):
    return templates.TemplateResponse(
        request,
        "bao_cao/index.html",
        {
            "request": request,
            "user": user, 
            "phan_loai_options": PHAN_LOAI_NGHE_NGHIEP,
            "loai_hinh_dac_thu": LOAI_HINH_DAC_THU
        },
    )


@router.get("/api/thong-ke")
def api_thong_ke(
    tu_ngay: Optional[str] = Query(None, description="YYYY-MM-DD"),
    den_ngay: Optional[str] = Query(None, description="YYYY-MM-DD"),
    phan_loai: Optional[str] = Query(None),
    loai_hs_dac_thu: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(require_login),
):
    try:
        dt_from, dt_to = _parse_dates(tu_ngay, den_ngay)
        return _query_stats(db, dt_from, dt_to, phan_loai, loai_hs_dac_thu)
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Lỗi api_thong_ke: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi thống kê: {exc}")


@router.get("/export-xlsx")
def export_xlsx(
    tu_ngay: Optional[str] = Query(None, description="YYYY-MM-DD"),
    den_ngay: Optional[str] = Query(None, description="YYYY-MM-DD"),
    phan_loai: Optional[str] = Query(None),
    loai_hs_dac_thu: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(require_login),
):
    try:
        dt_from, dt_to = _parse_dates(tu_ngay, den_ngay)
        stats = _query_stats(db, dt_from, dt_to, phan_loai, loai_hs_dac_thu)

        filter_lines: list[str] = []
        if tu_ngay:
            filter_lines.append(
                f"Từ ngày: {datetime.strptime(tu_ngay, '%Y-%m-%d').strftime('%d/%m/%Y')}"
            )
        if den_ngay:
            filter_lines.append(
                f"Đến ngày: {datetime.strptime(den_ngay, '%Y-%m-%d').strftime('%d/%m/%Y')}"
            )
        if phan_loai:
            filter_lines.append(f"Phân loại: {phan_loai}")
        if loai_hs_dac_thu:
            if loai_hs_dac_thu == "ALL":
                filter_lines.append("Hồ sơ đặc thù: Có")
            else:
                loai_hinh_str = LOAI_HINH_DAC_THU.get(loai_hs_dac_thu, loai_hs_dac_thu)
                filter_lines.append(f"Hồ sơ đặc thù: {loai_hinh_str}")

        if not filter_lines:
            filter_lines = ["Toàn bộ dữ liệu (không lọc)"]

        # ── Lấy danh sách chi tiết ─────────────────────────────────────────────
        conds = [DoiTuong.is_draft == False]
        if dt_from:
            conds.append(DoiTuong.created_at >= dt_from)
        if dt_to:
            conds.append(DoiTuong.created_at <= dt_to)
        if phan_loai:
            conds.append(DoiTuong.phan_loai_nghe_nghiep == phan_loai)
        if loai_hs_dac_thu:
            if loai_hs_dac_thu == "ALL":
                conds.append(DoiTuong.cccd.in_(select(HoSoDacThu.cccd)))
            else:
                conds.append(DoiTuong.cccd.in_(select(HoSoDacThu.cccd).where(HoSoDacThu.loai_hinh == loai_hs_dac_thu)))

        from sqlalchemy.orm import joinedload
        query = select(DoiTuong).where(and_(*conds)).options(
            joinedload(DoiTuong.lien_he),
            joinedload(DoiTuong.ho_so_dac_thu),
            joinedload(DoiTuong.nhan_than),
            joinedload(DoiTuong.phuong_tien),
            joinedload(DoiTuong.tai_chinh),
            joinedload(DoiTuong.qua_trinh),
        ).order_by(DoiTuong.created_at.desc())
        
        detailed_records = db.execute(query).unique().scalars().all()

        now = datetime.now()
        buf = _build_xlsx(stats, filter_lines, now, detailed_records)

        filename = f"bao-cao-thong-ke-{now.strftime('%Y%m%d-%H%M')}.xlsx"
        return StreamingResponse(
            buf,
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Lỗi export_xlsx: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi xuất Excel: {exc}")
