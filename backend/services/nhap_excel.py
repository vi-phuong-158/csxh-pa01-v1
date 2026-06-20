# -*- coding: utf-8 -*-
"""
Service Nhập Excel: sinh file mẫu nhiều sheet + import workbook.

File mẫu và importer dùng CHUNG định nghĩa SHEET_DEFS bên dưới —
header tiếng Việt cho cán bộ, importer tự ánh xạ về field kỹ thuật.
Thêm/sửa cột: chỉ sửa một chỗ này, hai phía tự khớp nhau.
"""
import io
import logging
import re
import unicodedata
import zipfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.orm import Session

from backend import constants
from backend.models.models import (
    AuditLog,
    DoiTuong,
    HoSoDacThu,
    LienHe,
    NhanThan,
    PhuongTien,
    QuaTrinhHoatDong,
    QuanHeDoiTuong,
    TaiChinh,
)
from backend.utils.text_utils import normalize_phone

logger = logging.getLogger(__name__)

# CCCD mẫu mới 12 số / CMND cũ 9 số — đồng bộ với backend/utils/validators.py
_CCCD_RE = re.compile(r"^(?:\d{9}|\d{12})$")

# Database Safety (CLAUDE.md): commit theo chunk tránh giữ write-lock lâu.
_CHUNK_SIZE = 100

# SQLite giới hạn ~999 biến bind / câu lệnh.
_IN_CLAUSE_SIZE = 500

# Số dòng được định dạng Text + gắn dropdown sẵn trong file mẫu.
_TEMPLATE_ROWS = 2000

# Giới hạn số lỗi hiển thị mỗi sheet (tránh trả HTML khổng lồ).
_MAX_ERRORS_SHOWN = 300

# Chống "zip bomb": .xlsx là file ZIP, 10MB nén có thể bung tới hàng GB XML khi
# pandas nạp toàn bộ sheet vào RAM. Chặn theo TỔNG kích thước bung (đọc từ
# metadata ZIP, không cần giải nén) trước khi đưa cho pandas.
_MAX_UNCOMPRESSED_MB = 200

# Cột "Hành động" sheet Đối tượng: THÊM MỚI (mặc định) hoặc CẬP NHẬT hồ sơ có sẵn.
_ACTION_ADD = "THÊM MỚI"
_ACTION_UPDATE = "CẬP NHẬT"

# Các field của DoiTuong được ghi đè ở chế độ CẬP NHẬT khi ô có giá trị
# (cccd là khóa, ho_ten/ngay_sinh xử lý riêng bên dưới).
_DOI_TUONG_OPT_FIELDS = (
    "gioi_tinh", "dan_toc", "ton_giao", "quoc_tich",
    "dia_chi_tinh", "dia_chi_xa", "que_quan", "noi_o_hien_nay",
    "phan_loai_nghe_nghiep", "chi_tiet_nghe_nghiep", "ghi_chu_chung",
)

# ============================================================
# ĐỊNH NGHĨA SHEET — nguồn sự thật duy nhất cho cả 2 chiều
# ============================================================
# Mỗi cột: label (header hiển thị), field (tên kỹ thuật),
# width, list (key dropdown, tùy chọn), text (định dạng Text, tùy chọn).

_DROPDOWNS = {
    "hanh_dong": ["THÊM MỚI", "CẬP NHẬT"],
    "gioi_tinh": constants.GIOI_TINH_OPTIONS,
    "dan_toc": constants.DANH_SACH_DAN_TOC,
    "ton_giao": constants.DANH_SACH_TON_GIAO,
    "quoc_gia": constants.DANH_SACH_QUOC_GIA,
    "tinh": constants.TINH_OPTIONS,
    "xa_phuong": constants.DANH_SACH_XA_PHU_THO,
    "nghe_nghiep": constants.PHAN_LOAI_NGHE_NGHIEP_OPTIONS,
    "loai_lien_he": constants.LOAI_LIEN_HE_OPTIONS,
    "ngan_hang": constants.DANH_SACH_NGAN_HANG,
    "loai_xe": constants.LOAI_XE_OPTIONS,
    "loai_quan_he": constants.LOAI_QUAN_HE,
    "loai_hinh_dac_thu": list(constants.LOAI_HINH_DAC_THU.values()),
}

SHEET_DEFS = {
    "doi_tuong": {
        "title": "1. Đối tượng",
        "aliases": {"đối tượng", "doi_tuong", "doi tuong"},
        "cols": [
            {"label": "Hành động", "field": "hanh_dong", "width": 12, "list": "hanh_dong"},
            {"label": "CCCD (*)", "field": "cccd", "width": 15, "text": True},
            {"label": "Họ và tên (*)", "field": "ho_ten", "width": 26},
            {"label": "Ngày sinh (dd/mm/yyyy)", "field": "ngay_sinh", "width": 16},
            {"label": "Giới tính", "field": "gioi_tinh", "width": 10, "list": "gioi_tinh"},
            {"label": "Dân tộc", "field": "dan_toc", "width": 10, "list": "dan_toc"},
            {"label": "Tôn giáo", "field": "ton_giao", "width": 11, "list": "ton_giao"},
            {"label": "Quốc tịch", "field": "quoc_tich", "width": 13, "list": "quoc_gia"},
            {"label": "Tỉnh/TP", "field": "dia_chi_tinh", "width": 10, "list": "tinh"},
            {"label": "Xã/Phường", "field": "dia_chi_xa", "width": 20, "list": "xa_phuong"},
            {"label": "Quê quán", "field": "que_quan", "width": 26},
            {"label": "Nơi ở hiện nay", "field": "noi_o_hien_nay", "width": 30},
            {"label": "Phân loại nghề nghiệp", "field": "phan_loai_nghe_nghiep", "width": 20, "list": "nghe_nghiep"},
            {"label": "Chi tiết nghề nghiệp", "field": "chi_tiet_nghe_nghiep", "width": 24},
            {"label": "Ghi chú chung", "field": "ghi_chu_chung", "width": 28},
        ],
    },
    "lien_he": {
        "title": "2. Liên hệ",
        "aliases": {"liên hệ", "lien_he", "lien he"},
        "cols": [
            {"label": "CCCD (*)", "field": "cccd", "width": 15, "text": True},
            {"label": "Loại liên hệ (*)", "field": "loai_lien_he", "width": 14, "list": "loai_lien_he"},
            {"label": "Giá trị (*)", "field": "gia_tri", "width": 24, "text": True},
            {"label": "Ghi chú", "field": "ghi_chu", "width": 30},
        ],
    },
    "tai_chinh": {
        "title": "3. Tài chính",
        "aliases": {"tài chính", "tai_chinh", "tai chinh"},
        "cols": [
            {"label": "CCCD (*)", "field": "cccd", "width": 15, "text": True},
            {"label": "Ngân hàng", "field": "ngan_hang", "width": 16, "list": "ngan_hang"},
            {"label": "Số tài khoản (*)", "field": "so_tai_khoan", "width": 20, "text": True},
            {"label": "Chủ tài khoản", "field": "chu_tai_khoan", "width": 22},
            {"label": "Ghi chú", "field": "ghi_chu", "width": 30},
        ],
    },
    "phuong_tien": {
        "title": "4. Phương tiện",
        "aliases": {"phương tiện", "phuong_tien", "phuong tien"},
        "cols": [
            {"label": "CCCD (*)", "field": "cccd", "width": 15, "text": True},
            {"label": "Loại xe", "field": "loai_xe", "width": 12, "list": "loai_xe"},
            {"label": "Biển kiểm soát (*)", "field": "bien_kiem_soat", "width": 16},
            {"label": "Tên phương tiện", "field": "ten_phuong_tien", "width": 22},
            {"label": "Ghi chú", "field": "ghi_chu", "width": 30},
        ],
    },
    "nhan_than": {
        "title": "5. Nhân thân",
        "aliases": {"nhân thân", "nhan_than", "nhan than", "quan hệ", "than_nhan"},
        "cols": [
            {"label": "CCCD (*)", "field": "cccd", "width": 15, "text": True},
            {"label": "Loại quan hệ (*)", "field": "loai_quan_he", "width": 14, "list": "loai_quan_he"},
            {"label": "Họ tên nhân thân (*)", "field": "ho_ten", "width": 24},
            {"label": "CCCD nhân thân", "field": "cccd_nhan_than", "width": 15, "text": True},
            {"label": "Ngày sinh (dd/mm/yyyy)", "field": "ngay_sinh", "width": 16},
            {"label": "Giới tính", "field": "gioi_tinh", "width": 10, "list": "gioi_tinh"},
            {"label": "Dân tộc", "field": "dan_toc", "width": 10, "list": "dan_toc"},
            {"label": "Tôn giáo", "field": "ton_giao", "width": 11, "list": "ton_giao"},
            {"label": "Quốc tịch", "field": "quoc_tich", "width": 13, "list": "quoc_gia"},
            {"label": "Nghề nghiệp", "field": "nghe_nghiep", "width": 20},
            {"label": "Nơi ở", "field": "noi_o", "width": 28},
            {"label": "Ghi chú", "field": "ghi_chu", "width": 24},
        ],
    },
    "qua_trinh": {
        "title": "6. Quá trình",
        "aliases": {"quá trình", "quá trình hoạt động", "qua_trinh", "qua trinh"},
        "cols": [
            {"label": "CCCD (*)", "field": "cccd", "width": 15, "text": True},
            {"label": "Từ ngày (dd/mm/yyyy)", "field": "ngay_bat_dau", "width": 15},
            {"label": "Đến ngày (dd/mm/yyyy)", "field": "ngay_ket_thuc", "width": 15},
            {"label": "Thời gian (chữ, nếu không rõ ngày)", "field": "thoi_gian", "width": 22},
            {"label": "Nội dung (*)", "field": "noi_dung", "width": 44},
            {"label": "Ghi chú", "field": "ghi_chu", "width": 24},
        ],
    },
    "dac_thu": {
        "title": "7. Hồ sơ đặc thù",
        "aliases": {"hồ sơ đặc thù", "hồ sơ csxh", "ho_so_dac_thu", "dac_thu"},
        "cols": [
            {"label": "CCCD (*)", "field": "cccd", "width": 15, "text": True},
            {"label": "Loại hình (*)", "field": "loai_hinh", "width": 38, "list": "loai_hinh_dac_thu"},
            {"label": "Nội dung chi tiết", "field": "noi_dung_chi_tiet", "width": 50},
            {"label": "Ghi chú", "field": "ghi_chu", "width": 24},
        ],
    },
}


# ============================================================
# SINH FILE MẪU
# ============================================================

_HEADER_FILL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN_BORDER = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

_HUONG_DAN_LINES = [
    ("HƯỚNG DẪN NHẬP LIỆU FILE EXCEL — VCFE DATABASE", True),
    ("", False),
    ("1. Chỉ sheet «1. Đối tượng» là BẮT BUỘC. Các sheet 2-7 dùng để bổ sung", False),
    ("   thông tin (mỗi người có thể có nhiều dòng SĐT, tài khoản, xe, nhân thân...).", False),
    ("2. Cột CCCD là khóa nối giữa các sheet: dòng ở sheet 2-7 chỉ được nhận khi", False),
    ("   CCCD đó có ở sheet 1 cùng file HOẶC đã tồn tại trong hệ thống.", False),
    ("   → Muốn BỔ SUNG thông tin cho hồ sơ đã có: bỏ trống sheet 1, điền sheet 2-7.", False),
    ("3. CCCD phải đúng 9 hoặc 12 chữ số. Cột CCCD đã định dạng kiểu Text sẵn —", False),
    ("   KHÔNG xóa cột hay copy đè định dạng khác, tránh Excel cắt mất số 0 đầu.", False),
    ("4. Ngày tháng nhập dạng dd/mm/yyyy (ví dụ 01/09/1990).", False),
    ("5. Các cột có mũi tên xổ xuống: chọn từ danh mục có sẵn, hạn chế gõ tay.", False),
    ("6. Cột đánh dấu (*) là bắt buộc — thiếu sẽ bị từ chối ở dòng đó.", False),
    ("7. Cột «Hành động» ở sheet «1. Đối tượng»:", False),
    ("   • THÊM MỚI (mặc định, để trống cũng được): tạo hồ sơ mới. Nếu CCCD đã có", False),
    ("     trong hệ thống sẽ bị từ chối để không ghi đè nhầm.", False),
    ("   • CẬP NHẬT: sửa hồ sơ CCCD đã có. CHỈ ô nào điền mới ghi đè, ô để trống", False),
    ("     giữ nguyên giá trị cũ. CCCD chưa có hồ sơ sẽ báo lỗi.", False),
    ("8. KHÔNG đổi tên sheet, không đổi dòng tiêu đề (dòng 1) của các sheet.", False),
    ("", False),
    ("VÍ DỤ ĐIỀN SHEET «1. Đối tượng»:", True),
    ("   CCCD: 025090001234 | Họ và tên: NGUYỄN VĂN A | Ngày sinh: 01/01/1990", False),
    ("   Giới tính: Nam | Quốc tịch: Việt Nam | Xã/Phường: Phường Việt Trì", False),
    ("", False),
    ("VÍ DỤ ĐIỀN SHEET «2. Liên hệ» (một người 2 số điện thoại = 2 dòng):", True),
    ("   025090001234 | SĐT | 0912345678 | SĐT chính", False),
    ("   025090001234 | SĐT | 0987654321 | SĐT phụ", False),
    ("", False),
    ("Sau khi điền xong: lưu file rồi tải lên tại trang «Nhập dữ liệu từ Excel».", False),
    ("Hệ thống sẽ báo kết quả theo từng sheet, dòng nào lỗi sẽ ghi rõ lý do.", False),
]


def _write_huong_dan(ws):
    ws.title = "HƯỚNG DẪN"
    ws.column_dimensions["A"].width = 100
    for i, (line, bold) in enumerate(_HUONG_DAN_LINES, 1):
        cell = ws.cell(row=i, column=1, value=line)
        if bold:
            cell.font = Font(bold=True, color="0F766E", size=12)
    ws.sheet_view.showGridLines = False


def _write_danh_muc(wb) -> dict:
    """Sheet ẩn chứa danh mục cho dropdown (Data Validation tham chiếu range —
    danh sách dài quá 255 ký tự không nhét trực tiếp vào formula được)."""
    dm = wb.create_sheet("DM")
    dm.sheet_state = "hidden"
    refs = {}
    for col_idx, (key, values) in enumerate(_DROPDOWNS.items(), 1):
        letter = get_column_letter(col_idx)
        dm.cell(row=1, column=col_idx, value=key)
        for row_idx, val in enumerate(values, 2):
            dm.cell(row=row_idx, column=col_idx, value=val)
        refs[key] = f"DM!${letter}$2:${letter}${len(values) + 1}"
    return refs


def _write_data_sheet(wb, sheet_def: dict, dm_refs: dict):
    ws = wb.create_sheet(sheet_def["title"])
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(sheet_def["cols"], 1):
        letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=col["label"])
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER
        ws.column_dimensions[letter].width = col["width"]

        # Định dạng Text chống Excel cắt số 0 đầu của CCCD/SĐT/số TK.
        if col.get("text"):
            for r in range(2, _TEMPLATE_ROWS + 2):
                ws.cell(row=r, column=col_idx).number_format = "@"

        if col.get("list"):
            dv = DataValidation(
                type="list", formula1=dm_refs[col["list"]], allow_blank=True,
                showErrorMessage=True, errorTitle="Giá trị không hợp lệ",
                error="Vui lòng chọn từ danh mục xổ xuống.",
            )
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}{_TEMPLATE_ROWS + 1}")


def build_template() -> bytes:
    """Sinh file mẫu mau_ho_so_csxh.xlsx: 1 sheet hướng dẫn + 7 sheet dữ liệu."""
    wb = Workbook()
    _write_huong_dan(wb.active)
    dm_refs = _write_danh_muc(wb)
    for sheet_def in SHEET_DEFS.values():
        _write_data_sheet(wb, sheet_def, dm_refs)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# IMPORT WORKBOOK
# ============================================================

def _norm_text(s: str) -> str:
    """Chuẩn hóa header/tên sheet để so khớp: bỏ (*), bỏ số thứ tự, gọn khoảng trắng."""
    s = unicodedata.normalize("NFC", str(s))
    s = re.sub(r"\([^)]*\)", "", s)          # bỏ chú thích trong ngoặc
    s = re.sub(r"^\s*\d+\s*[\.\)]\s*", "", s)  # bỏ "1. " đầu tên sheet
    return re.sub(r"\s+", " ", s).strip().casefold()


def _norm_action(raw: str) -> str:
    """Chuẩn hóa cột Hành động -> CẬP NHẬT hoặc THÊM MỚI (mặc định khi trống).
    Chịu được gõ tay không dấu/hoa thường ('cap nhat', 'update', 'sửa')."""
    s = unicodedata.normalize("NFD", str(raw or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s).strip().casefold()
    if s in ("cap nhat", "capnhat", "update", "sua", "ghi de"):
        return _ACTION_UPDATE
    return _ACTION_ADD


def _build_colmap(df: pd.DataFrame, cols: list) -> dict:
    """Map field kỹ thuật -> tên cột gốc trong DataFrame. Nhận cả header
    tiếng Việt của file mẫu lẫn tên field snake_case (file cũ)."""
    wanted = {}
    for col in cols:
        wanted[_norm_text(col["label"])] = col["field"]
        wanted[col["field"]] = col["field"]
    out = {}
    for orig in df.columns:
        field = wanted.get(_norm_text(orig))
        if field and field not in out:
            out[field] = orig
    return out


def _fix_cccd(v: str) -> str:
    """Sửa CCCD bị Excel cắt số 0 đầu (lưu dạng số): 10-11 chữ số -> đệm về 12,
    8 chữ số -> đệm về 9. File mẫu mới định dạng Text nên hiếm khi cần."""
    v = re.sub(r"[\s.]", "", v)
    if v.endswith(".0"):
        v = v[:-2]
    if v.isdigit():
        if len(v) in (10, 11):
            return v.zfill(12)
        if len(v) == 8:
            return v.zfill(9)
    return v


def _parse_date(raw: str):
    """-> (date | None, ok). Chuỗi rỗng coi là hợp lệ (None)."""
    if not raw:
        return None, True
    try:
        return pd.to_datetime(raw, dayfirst=True).date(), True
    except Exception:
        return None, False


def _load_existing_cccd(db: Session, cccds: set) -> set:
    found = set()
    items = list(cccds)
    for i in range(0, len(items), _IN_CLAUSE_SIZE):
        batch = items[i:i + _IN_CLAUSE_SIZE]
        rows = db.execute(select(DoiTuong.cccd).where(DoiTuong.cccd.in_(batch))).all()
        found.update(r[0] for r in rows)
    return found


def _load_satellite_keys(db: Session, model_cols, cccds: set) -> set:
    """Load key chống trùng của bảng vệ tinh cho các CCCD liên quan.
    model_cols: tuple cột SQLAlchemy, cột đầu phải là cccd."""
    keys = set()
    items = list(cccds)
    for i in range(0, len(items), _IN_CLAUSE_SIZE):
        batch = items[i:i + _IN_CLAUSE_SIZE]
        rows = db.execute(select(*model_cols).where(model_cols[0].in_(batch))).all()
        keys.update(tuple(r) for r in rows)
    return keys


def _load_genders(db: Session, cccds: set) -> dict:
    """{cccd: gioi_tinh} cho các ĐT chính — định hướng quan hệ Cha/Mẹ-Con."""
    out = {}
    items = list(cccds)
    for i in range(0, len(items), _IN_CLAUSE_SIZE):
        batch = items[i:i + _IN_CLAUSE_SIZE]
        rows = db.execute(select(DoiTuong.cccd, DoiTuong.gioi_tinh).where(DoiTuong.cccd.in_(batch))).all()
        out.update({r[0]: r[1] for r in rows})
    return out


def _load_edges_touching(db: Session, cccds: set) -> set:
    """Các cạnh (cccd_1, cccd_2, loai_quan_he) đã tồn tại có dính tới cccds.
    Mọi cạnh sinh từ nhân thân đều dính 1 ĐT chính, nên set này đủ để chống
    trùng (không vi phạm unique index uq_quan_he_cap -> tránh IntegrityError)."""
    edges = set()
    items = list(cccds)
    for i in range(0, len(items), _IN_CLAUSE_SIZE):
        batch = items[i:i + _IN_CLAUSE_SIZE]
        rows = db.execute(
            select(QuanHeDoiTuong.cccd_1, QuanHeDoiTuong.cccd_2, QuanHeDoiTuong.loai_quan_he)
            .where(QuanHeDoiTuong.cccd_1.in_(batch) | QuanHeDoiTuong.cccd_2.in_(batch))
        ).all()
        edges.update(tuple(r) for r in rows)
    return edges


class _SheetReport:
    def __init__(self, title: str):
        self.title = title
        self.total = 0
        self.success = 0
        self.errors = []

    def err(self, row_num, msg):
        self.errors.append({"row": row_num, "msg": msg})

    @property
    def failed(self):
        return len(self.errors)

    def as_dict(self):
        errors = self.errors
        if len(errors) > _MAX_ERRORS_SHOWN:
            hidden = len(errors) - _MAX_ERRORS_SHOWN
            errors = errors[:_MAX_ERRORS_SHOWN] + [
                {"row": "…", "msg": f"(và {hidden} lỗi khác — sửa các lỗi trên rồi import lại)"}
            ]
        return {"title": self.title, "total": self.total,
                "success": self.success, "failed": self.failed, "errors": errors}


def _iter_rows(df: pd.DataFrame, colmap: dict):
    """Sinh (row_num_excel, getter). getter(field) -> str đã strip, '' nếu trống.
    Bỏ qua dòng trống hoàn toàn."""
    for idx, row in df.iterrows():
        def get(field, _row=row):
            orig = colmap.get(field)
            if orig is None:
                return ""
            v = _row[orig]
            if pd.isna(v):
                return ""
            return str(v).strip()

        if any(get(f) for f in colmap):
            yield idx + 2, get


def _find_sheets(sheets_raw: dict) -> dict:
    """Khớp tên sheet trong workbook với SHEET_DEFS (chịu được sai khác nhỏ)."""
    found = {}
    for key, sdef in SHEET_DEFS.items():
        for name, df in sheets_raw.items():
            if _norm_text(name) in sdef["aliases"]:
                found[key] = df
                break
    # File cũ 1 sheet duy nhất với header snake_case -> coi là sheet Đối tượng
    # (chỉ khi KHÔNG nhận diện được sheet nào — tránh nuốt nhầm sheet vệ tinh
    # khi cán bộ upload file chỉ có mỗi sheet Liên hệ chẳng hạn).
    if not found and len(sheets_raw) == 1:
        found["doi_tuong"] = next(iter(sheets_raw.values()))
    return found


def _check_xlsx_bomb(contents: bytes) -> bool:
    """True nếu tổng kích thước bung của workbook vượt ngưỡng an toàn.
    Đọc field uncompressed-size trong central directory của ZIP — không giải nén,
    nên rẻ và an toàn (zip bomb tự khai báo kích thước lớn ở đây)."""
    try:
        with zipfile.ZipFile(io.BytesIO(contents)) as zf:
            total = sum(zi.file_size for zi in zf.infolist())
    except zipfile.BadZipFile:
        return False  # không phải ZIP (vd .xls cũ) — để pandas tự xử lý/raise
    return total > _MAX_UNCOMPRESSED_MB * 1024 * 1024


def import_workbook(db: Session, contents: bytes, username: str, filename: str) -> dict:
    """Import toàn bộ workbook. Trả về report cho template kết quả.

    Lưu ý: import là PARTIAL-COMMIT (commit theo chunk _CHUNK_SIZE để tránh giữ
    write-lock lâu — yêu cầu CLAUDE.md), KHÔNG nguyên tử. Nếu lỗi hạ tầng xảy ra
    giữa chừng, các chunk đã commit vẫn còn; lỗi DỮ LIỆU thì được bắt theo từng
    dòng nên không làm hỏng cả lượt import."""
    if _check_xlsx_bomb(contents):
        return {"error": f"File giải nén quá lớn (> {_MAX_UNCOMPRESSED_MB}MB) — "
                         "nghi ngờ file hỏng hoặc độc hại, đã từ chối xử lý."}
    sheets_raw = pd.read_excel(io.BytesIO(contents), sheet_name=None, dtype=str)
    found = _find_sheets(sheets_raw)

    if not found:
        return {"error": "Không nhận diện được sheet nào trong file. "
                         "Vui lòng dùng đúng file mẫu (không đổi tên sheet)."}

    reports = []
    known_cccd: set[str] = set()  # CCCD chắc chắn tồn tại (DB hoặc vừa nhập)

    # ── Sheet 1: Đối tượng ────────────────────────────────────────────
    if "doi_tuong" in found:
        rep = _import_doi_tuong(db, found["doi_tuong"], known_cccd)
        reports.append(rep)

    # ── Sheet vệ tinh ─────────────────────────────────────────────────
    satellite_handlers = [
        ("lien_he", _import_lien_he),
        ("tai_chinh", _import_tai_chinh),
        ("phuong_tien", _import_phuong_tien),
        ("nhan_than", _import_nhan_than),
        ("qua_trinh", _import_qua_trinh),
        ("dac_thu", _import_dac_thu),
    ]
    for key, handler in satellite_handlers:
        if key in found:
            rep = handler(db, found[key], known_cccd)
            reports.append(rep)

    total_success = sum(r.success for r in reports)
    total_failed = sum(r.failed for r in reports)

    # Audit log: import hàng loạt là thao tác nhạy cảm, phải trace được.
    try:
        db.add(AuditLog(
            bang="doi_tuong",
            hanh_dong="BULK_IMPORT",
            khoa_chinh=filename or "excel",
            du_lieu_moi=f"Import Excel: {total_success} thành công, {total_failed} lỗi "
                        f"({len(reports)} sheet)",
            nguoi_thuc_hien=username,
        ))
        db.commit()
    except Exception:
        logger.exception("Không thể ghi audit log BULK_IMPORT")

    return {
        "success": total_success,
        "failed": total_failed,
        "sheets": [r.as_dict() for r in reports],
    }


def _validate_cccd_ref(cccd_raw: str, known_cccd: set, rep: _SheetReport, row_num: int):
    """Validate CCCD tham chiếu ở sheet vệ tinh. -> cccd hợp lệ hoặc None."""
    cccd = _fix_cccd(cccd_raw)
    if not _CCCD_RE.fullmatch(cccd):
        rep.err(row_num, "CCCD không hợp lệ (phải đúng 9 hoặc 12 chữ số)")
        return None
    if cccd not in known_cccd:
        rep.err(row_num, f"CCCD {cccd} chưa có hồ sơ (không có ở sheet Đối tượng và chưa tồn tại trong hệ thống)")
        return None
    return cccd


def _resolve_satellite_refs(db: Session, df: pd.DataFrame, colmap: dict, known_cccd: set):
    """Bổ sung vào known_cccd các CCCD tham chiếu đã tồn tại sẵn trong DB
    (chế độ bổ sung thông tin cho hồ sơ có sẵn)."""
    refs = set()
    for _, get in _iter_rows(df, colmap):
        c = _fix_cccd(get("cccd"))
        if _CCCD_RE.fullmatch(c) and c not in known_cccd:
            refs.add(c)
    if refs:
        known_cccd.update(_load_existing_cccd(db, refs))


def _import_doi_tuong(db: Session, df: pd.DataFrame, known_cccd: set) -> _SheetReport:
    sdef = SHEET_DEFS["doi_tuong"]
    rep = _SheetReport(sdef["title"])
    colmap = _build_colmap(df, sdef["cols"])

    if "cccd" not in colmap or "ho_ten" not in colmap:
        rep.err(1, "Sheet thiếu cột bắt buộc CCCD / Họ và tên — dùng đúng file mẫu")
        return rep

    # Batch check trùng hệ thống: 1 lượt IN(...) cho toàn bộ file (tránh N+1).
    all_cccd = set()
    for _, get in _iter_rows(df, colmap):
        c = _fix_cccd(get("cccd"))
        if _CCCD_RE.fullmatch(c):
            all_cccd.add(c)
    existing = _load_existing_cccd(db, all_cccd)
    known_cccd.update(existing)

    seen_in_file: set[str] = set()
    pending = 0

    for row_num, get in _iter_rows(df, colmap):
        rep.total += 1
        cccd = _fix_cccd(get("cccd"))
        ho_ten = get("ho_ten")
        action = _norm_action(get("hanh_dong"))

        if not cccd:
            rep.err(row_num, "CCCD bị trống")
            continue
        if not _CCCD_RE.fullmatch(cccd):
            rep.err(row_num, "CCCD không hợp lệ (phải đúng 9 hoặc 12 chữ số)")
            continue
        if cccd in seen_in_file:
            rep.err(row_num, f"CCCD {cccd} bị lặp trong chính file Excel")
            continue

        ngay_sinh, ok = _parse_date(get("ngay_sinh"))
        if not ok:
            rep.err(row_num, f"Ngày sinh không hợp lệ ({get('ngay_sinh')!r}) — cần dd/mm/yyyy")
            continue

        if action == _ACTION_UPDATE:
            # Chỉ ghi đè các ô có giá trị; ô trống giữ nguyên dữ liệu cũ.
            obj = db.get(DoiTuong, cccd) if cccd in existing else None
            if obj is None:
                rep.err(row_num, f"CCCD {cccd} chưa có hồ sơ để cập nhật — dùng «THÊM MỚI»")
                continue
            if ho_ten:
                obj.ho_ten = ho_ten.upper()
            if get("ngay_sinh"):
                obj.ngay_sinh = ngay_sinh
            for field in _DOI_TUONG_OPT_FIELDS:
                val = get(field)
                if val:
                    setattr(obj, field, val)
        else:
            if not ho_ten:
                rep.err(row_num, "Họ tên bị trống")
                continue
            if cccd in existing:
                rep.err(row_num, f"CCCD {cccd} đã tồn tại trong hệ thống — dùng «CẬP NHẬT» nếu muốn sửa")
                continue
            db.add(DoiTuong(
                cccd=cccd,
                ho_ten=ho_ten.upper(),
                ngay_sinh=ngay_sinh,
                gioi_tinh=get("gioi_tinh") or None,
                dan_toc=get("dan_toc") or None,
                ton_giao=get("ton_giao") or None,
                quoc_tich=get("quoc_tich") or None,
                dia_chi_tinh=get("dia_chi_tinh") or None,
                dia_chi_xa=get("dia_chi_xa") or None,
                que_quan=get("que_quan") or None,
                noi_o_hien_nay=get("noi_o_hien_nay") or None,
                phan_loai_nghe_nghiep=get("phan_loai_nghe_nghiep") or None,
                chi_tiet_nghe_nghiep=get("chi_tiet_nghe_nghiep") or None,
                ghi_chu_chung=get("ghi_chu_chung") or None,
                is_draft=False,
            ))
            known_cccd.add(cccd)

        seen_in_file.add(cccd)
        rep.success += 1
        pending += 1
        if pending >= _CHUNK_SIZE:
            db.commit()
            pending = 0

    if pending:
        db.commit()
    return rep


def _import_lien_he(db: Session, df: pd.DataFrame, known_cccd: set) -> _SheetReport:
    sdef = SHEET_DEFS["lien_he"]
    rep = _SheetReport(sdef["title"])
    colmap = _build_colmap(df, sdef["cols"])
    _resolve_satellite_refs(db, df, colmap, known_cccd)

    # Chống nhập trùng khi upload lại cùng file.
    existing_keys = _load_satellite_keys(
        db, (LienHe.cccd, LienHe.loai_lien_he, LienHe.gia_tri), known_cccd)
    pending = 0

    for row_num, get in _iter_rows(df, colmap):
        rep.total += 1
        cccd = _validate_cccd_ref(get("cccd"), known_cccd, rep, row_num)
        if not cccd:
            continue
        loai = get("loai_lien_he")
        gia_tri = get("gia_tri")
        if not loai or not gia_tri:
            rep.err(row_num, "Thiếu Loại liên hệ hoặc Giá trị")
            continue
        if loai == "SĐT":
            gia_tri = normalize_phone(gia_tri)
        key = (cccd, loai, gia_tri)
        if key in existing_keys:
            rep.err(row_num, f"Liên hệ {loai} {gia_tri} của CCCD {cccd} đã có")
            continue
        existing_keys.add(key)
        db.add(LienHe(cccd=cccd, loai_lien_he=loai, gia_tri=gia_tri,
                      ghi_chu=get("ghi_chu") or None))
        rep.success += 1
        pending += 1
        if pending >= _CHUNK_SIZE:
            db.commit()
            pending = 0

    if pending:
        db.commit()
    return rep


def _import_tai_chinh(db: Session, df: pd.DataFrame, known_cccd: set) -> _SheetReport:
    sdef = SHEET_DEFS["tai_chinh"]
    rep = _SheetReport(sdef["title"])
    colmap = _build_colmap(df, sdef["cols"])
    _resolve_satellite_refs(db, df, colmap, known_cccd)

    existing_keys = _load_satellite_keys(
        db, (TaiChinh.cccd, TaiChinh.so_tai_khoan), known_cccd)
    pending = 0

    for row_num, get in _iter_rows(df, colmap):
        rep.total += 1
        cccd = _validate_cccd_ref(get("cccd"), known_cccd, rep, row_num)
        if not cccd:
            continue
        so_tk = re.sub(r"[\s.\-]", "", get("so_tai_khoan"))
        if so_tk.endswith(".0"):
            so_tk = so_tk[:-2]
        if not so_tk:
            rep.err(row_num, "Thiếu Số tài khoản")
            continue
        key = (cccd, so_tk)
        if key in existing_keys:
            rep.err(row_num, f"Số tài khoản {so_tk} của CCCD {cccd} đã có")
            continue
        existing_keys.add(key)
        db.add(TaiChinh(cccd=cccd, ngan_hang=get("ngan_hang") or None,
                        so_tai_khoan=so_tk, chu_tai_khoan=get("chu_tai_khoan") or None,
                        ghi_chu=get("ghi_chu") or None))
        rep.success += 1
        pending += 1
        if pending >= _CHUNK_SIZE:
            db.commit()
            pending = 0

    if pending:
        db.commit()
    return rep


def _import_phuong_tien(db: Session, df: pd.DataFrame, known_cccd: set) -> _SheetReport:
    sdef = SHEET_DEFS["phuong_tien"]
    rep = _SheetReport(sdef["title"])
    colmap = _build_colmap(df, sdef["cols"])
    _resolve_satellite_refs(db, df, colmap, known_cccd)

    existing_raw = _load_satellite_keys(
        db, (PhuongTien.cccd, PhuongTien.bien_kiem_soat), known_cccd)
    # So khớp biển số sau khi bỏ khoảng trắng/chấm/gạch + viết hoa.
    norm_bks = lambda s: re.sub(r"[\s.\-]", "", (s or "")).upper()
    existing_keys = {(c, norm_bks(b)) for c, b in existing_raw}
    pending = 0

    for row_num, get in _iter_rows(df, colmap):
        rep.total += 1
        cccd = _validate_cccd_ref(get("cccd"), known_cccd, rep, row_num)
        if not cccd:
            continue
        bks = get("bien_kiem_soat")
        if not bks:
            rep.err(row_num, "Thiếu Biển kiểm soát")
            continue
        key = (cccd, norm_bks(bks))
        if key in existing_keys:
            rep.err(row_num, f"Biển kiểm soát {bks} của CCCD {cccd} đã có")
            continue
        existing_keys.add(key)
        db.add(PhuongTien(cccd=cccd, loai_xe=get("loai_xe") or None,
                          bien_kiem_soat=bks.upper(),
                          ten_phuong_tien=get("ten_phuong_tien") or None,
                          ghi_chu=get("ghi_chu") or None))
        rep.success += 1
        pending += 1
        if pending >= _CHUNK_SIZE:
            db.commit()
            pending = 0

    if pending:
        db.commit()
    return rep


# Ánh xạ nhãn nhân thân (LOAI_QUAN_HE) -> (key graph, vai trò của nhân thân).
# role: "cha_me" = nhân thân là cha/mẹ của ĐT chính; "con" = nhân thân là con
# của ĐT chính (key tính theo giới tính ĐT chính); "doi_xung" = quan hệ ngang hàng.
_NHAN_THAN_GRAPH = {
    "Bố":          ("Cha-Con", "cha_me"),
    "Cha":         ("Cha-Con", "cha_me"),
    "Mẹ":          ("Mẹ-Con",  "cha_me"),
    "Con trai":    (None,       "con"),
    "Con gái":     (None,       "con"),
    "Con":         (None,       "con"),
    "Vợ":          ("Vợ chồng", "doi_xung"),
    "Chồng":       ("Vợ chồng", "doi_xung"),
    "Anh":         ("Anh chị em", "doi_xung"),
    "Chị":         ("Anh chị em", "doi_xung"),
    "Em trai":     ("Anh chị em", "doi_xung"),
    "Em gái":      ("Anh chị em", "doi_xung"),
    "Bạn bè":      ("Bạn bè", "doi_xung"),
    "Đồng nghiệp": ("Đồng nghiệp", "doi_xung"),
}


def _build_quan_he_edge(main_cccd: str, nt_cccd: str, loai_nhan_than: str, main_gender):
    """Dựng cạnh (cccd_1, cccd_2, loai_quan_he) cho graph từ một dòng nhân thân.
    Cha/mẹ đứng cccd_1, con đứng cccd_2; quan hệ đối xứng chuẩn hóa theo min/max
    để dedup ổn định. Nhãn lạ -> 'Khác' (đối xứng)."""
    key, role = _NHAN_THAN_GRAPH.get(loai_nhan_than, ("Khác", "doi_xung"))
    if role == "cha_me":
        return (nt_cccd, main_cccd, key)            # nhân thân (cha/mẹ) -> con là ĐT chính
    if role == "con":
        k = "Mẹ-Con" if (main_gender or "").strip() == "Nữ" else "Cha-Con"
        return (main_cccd, nt_cccd, k)              # ĐT chính (cha/mẹ) -> con là nhân thân
    c1, c2 = sorted((main_cccd, nt_cccd))
    return (c1, c2, key)


def _import_nhan_than(db: Session, df: pd.DataFrame, known_cccd: set) -> _SheetReport:
    sdef = SHEET_DEFS["nhan_than"]
    rep = _SheetReport(sdef["title"])
    colmap = _build_colmap(df, sdef["cols"])
    _resolve_satellite_refs(db, df, colmap, known_cccd)

    # Preload chống N+1: giới tính ĐT chính (định hướng cha/mẹ-con) và các cạnh
    # quan hệ đã tồn tại (chống trùng -> không vi phạm unique index).
    main_cccds = {c for _, get in _iter_rows(df, colmap)
                  if _CCCD_RE.fullmatch(c := _fix_cccd(get("cccd")))}
    gender = _load_genders(db, main_cccds)
    existing_edges = _load_edges_touching(db, main_cccds)
    # Chống nhập trùng bản ghi nhân thân: cùng (CCCD, loại quan hệ, họ tên).
    existing_keys = _load_satellite_keys(
        db, (NhanThan.cccd, NhanThan.loai_quan_he, NhanThan.ho_ten), known_cccd)
    seen_edges: set = set()

    pending = 0
    edges_added = 0
    drafts_added = 0

    for row_num, get in _iter_rows(df, colmap):
        rep.total += 1
        cccd = _validate_cccd_ref(get("cccd"), known_cccd, rep, row_num)
        if not cccd:
            continue
        loai = get("loai_quan_he")
        ho_ten = get("ho_ten")
        if not loai or not ho_ten:
            rep.err(row_num, "Thiếu Loại quan hệ hoặc Họ tên nhân thân")
            continue
        ngay_sinh, ok = _parse_date(get("ngay_sinh"))
        if not ok:
            rep.err(row_num, f"Ngày sinh không hợp lệ ({get('ngay_sinh')!r}) — cần dd/mm/yyyy")
            continue
        nt_key = (cccd, loai, ho_ten.upper())
        if nt_key in existing_keys:
            rep.err(row_num, "Nhân thân trùng (cùng CCCD/loại quan hệ/họ tên) đã có")
            continue
        existing_keys.add(nt_key)
        cccd_nt = _fix_cccd(get("cccd_nhan_than")) if get("cccd_nhan_than") else None
        db.add(NhanThan(
            cccd=cccd, loai_quan_he=loai, ho_ten=ho_ten.upper(),
            cccd_nhan_than=cccd_nt, ngay_sinh=ngay_sinh,
            gioi_tinh=get("gioi_tinh") or None,
            dan_toc=get("dan_toc") or None,
            ton_giao=get("ton_giao") or None,
            quoc_tich=get("quoc_tich") or None,
            nghe_nghiep=get("nghe_nghiep") or None,
            noi_o=get("noi_o") or None,
            ghi_chu=get("ghi_chu") or None,
        ))
        rep.success += 1
        pending += 1

        # Graph: nhân thân có CCCD hợp lệ và khác ĐT chính -> tạo cạnh quan hệ.
        if cccd_nt and _CCCD_RE.fullmatch(cccd_nt) and cccd_nt != cccd:
            # Bảo đảm nhân thân tồn tại như một hồ sơ (ràng buộc FK). Chưa có ->
            # tạo hồ sơ nháp để cán bộ bổ sung sau. SQLAlchemy flush cha trước con.
            if cccd_nt not in known_cccd:
                db.add(DoiTuong(cccd=cccd_nt, ho_ten=ho_ten.upper(), is_draft=True))
                known_cccd.add(cccd_nt)
                drafts_added += 1
                pending += 1
            edge = _build_quan_he_edge(cccd, cccd_nt, loai, gender.get(cccd))
            if edge not in existing_edges and edge not in seen_edges:
                seen_edges.add(edge)
                c1, c2, lqh = edge
                db.add(QuanHeDoiTuong(cccd_1=c1, cccd_2=c2, loai_quan_he=lqh,
                                      mo_ta=get("ghi_chu") or None))
                edges_added += 1
                pending += 1

        if pending >= _CHUNK_SIZE:
            db.commit()
            pending = 0

    if pending:
        db.commit()
    if edges_added or drafts_added:
        logger.info("Nhập Excel nhân thân: tạo %d cạnh quan hệ, %d hồ sơ nháp",
                    edges_added, drafts_added)
    return rep


def _import_qua_trinh(db: Session, df: pd.DataFrame, known_cccd: set) -> _SheetReport:
    sdef = SHEET_DEFS["qua_trinh"]
    rep = _SheetReport(sdef["title"])
    colmap = _build_colmap(df, sdef["cols"])
    _resolve_satellite_refs(db, df, colmap, known_cccd)

    # Chống nhập trùng khi upload lại: cùng (CCCD, nội dung, ngày bắt đầu).
    existing_keys = _load_satellite_keys(
        db, (QuaTrinhHoatDong.cccd, QuaTrinhHoatDong.noi_dung, QuaTrinhHoatDong.ngay_bat_dau), known_cccd)
    pending = 0

    for row_num, get in _iter_rows(df, colmap):
        rep.total += 1
        cccd = _validate_cccd_ref(get("cccd"), known_cccd, rep, row_num)
        if not cccd:
            continue
        noi_dung = get("noi_dung")
        if not noi_dung:
            rep.err(row_num, "Thiếu Nội dung")
            continue
        ngay_bd, ok1 = _parse_date(get("ngay_bat_dau"))
        ngay_kt, ok2 = _parse_date(get("ngay_ket_thuc"))
        if not ok1 or not ok2:
            rep.err(row_num, "Từ ngày / Đến ngày không hợp lệ — cần dd/mm/yyyy")
            continue
        qt_key = (cccd, noi_dung, ngay_bd)
        if qt_key in existing_keys:
            rep.err(row_num, "Quá trình trùng (cùng CCCD/nội dung/ngày bắt đầu) đã có")
            continue
        existing_keys.add(qt_key)
        db.add(QuaTrinhHoatDong(
            cccd=cccd, ngay_bat_dau=ngay_bd, ngay_ket_thuc=ngay_kt,
            thoi_gian=get("thoi_gian") or None, noi_dung=noi_dung,
            ghi_chu=get("ghi_chu") or None,
        ))
        rep.success += 1
        pending += 1
        if pending >= _CHUNK_SIZE:
            db.commit()
            pending = 0

    if pending:
        db.commit()
    return rep


# Nhãn hiển thị -> key lưu DB (UI lọc theo key, ví dụ "Hon_Nhan_NN").
_DAC_THU_LABEL_TO_KEY = {v: k for k, v in constants.LOAI_HINH_DAC_THU.items()}


def _import_dac_thu(db: Session, df: pd.DataFrame, known_cccd: set) -> _SheetReport:
    sdef = SHEET_DEFS["dac_thu"]
    rep = _SheetReport(sdef["title"])
    colmap = _build_colmap(df, sdef["cols"])
    _resolve_satellite_refs(db, df, colmap, known_cccd)

    # Chống nhập trùng khi upload lại: cùng (CCCD, loại hình, nội dung chi tiết).
    existing_keys = _load_satellite_keys(
        db, (HoSoDacThu.cccd, HoSoDacThu.loai_hinh, HoSoDacThu.noi_dung_chi_tiet), known_cccd)
    pending = 0

    for row_num, get in _iter_rows(df, colmap):
        rep.total += 1
        cccd = _validate_cccd_ref(get("cccd"), known_cccd, rep, row_num)
        if not cccd:
            continue
        loai_raw = get("loai_hinh")
        loai = _DAC_THU_LABEL_TO_KEY.get(loai_raw, loai_raw)
        if loai not in constants.LOAI_HINH_DAC_THU:
            rep.err(row_num, f"Loại hình không hợp lệ ({loai_raw!r}) — chọn từ danh mục xổ xuống")
            continue
        noi_dung_ct = get("noi_dung_chi_tiet") or None
        dt_key = (cccd, loai, noi_dung_ct)
        if dt_key in existing_keys:
            rep.err(row_num, "Hồ sơ đặc thù trùng (cùng CCCD/loại hình/nội dung) đã có")
            continue
        existing_keys.add(dt_key)
        db.add(HoSoDacThu(
            cccd=cccd, loai_hinh=loai,
            noi_dung_chi_tiet=noi_dung_ct,
            ghi_chu=get("ghi_chu") or None,
        ))
        rep.success += 1
        pending += 1
        if pending >= _CHUNK_SIZE:
            db.commit()
            pending = 0

    if pending:
        db.commit()
    return rep
