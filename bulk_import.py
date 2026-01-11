# -*- coding: utf-8 -*-
"""
Module Bulk Import cho Security Profile 360
Xử lý nhập liệu hàng loạt từ file Excel đa sheet
"""

import pandas as pd
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from database import get_connection
from constants import (
    DANH_SACH_XA_PHU_THO, GIOI_TINH_OPTIONS, TINH_OPTIONS, 
    PHAN_LOAI_NGHE_NGHIEP_OPTIONS, LOAI_LIEN_HE_OPTIONS, LOAI_XE_OPTIONS,
    DANH_SACH_QUOC_GIA, LOAI_HINH_TO_CHUC_NN, HINH_THUC_DU_HOC, 
    KET_QUA_XAC_MINH, DANH_SACH_NGAN_HANG
)


# ============================================
# TẠO FILE EXCEL MẪU 5 SHEET
# ============================================

# Định nghĩa cấu trúc cột cho từng loại CSXH
CSXH_TEMPLATES = {
    "Hon_Nhan_NN": {
        "name": "Hôn nhân với người nước ngoài",
        "headers": ["CCCD (*)", "Quốc tịch người nước ngoài (*)", "Họ tên người nước ngoài (*)", 
                   "Năm kết hôn", "Nơi đăng ký kết hôn", "Tình trạng hiện tại", 
                   "Địa chỉ hiện tại", "Ghi chú"],
        "sample": ["001234567890", "Trung Quốc", "WANG Xiaoming", 
                   "2020", "UBND TP Việt Trì", "Đang chung sống", 
                   "Xã Thanh Ba, huyện Thanh Ba, Phú Thọ", "Đang theo dõi"]
    },
    "Lam_Viec_NN": {
        "name": "Làm việc cho tổ chức nước ngoài",
        "headers": ["CCCD (*)", "Tên tổ chức (*)", "Quốc gia gốc của tổ chức (*)", 
                   "Loại hình (FDI/NGO/Khác)", "Vị trí công việc", 
                   "Năm bắt đầu", "Năm kết thúc", "Địa chỉ làm việc", "Ghi chú"],
        "sample": ["001234567890", "Samsung Electronics Vietnam", "Hàn Quốc", 
                   "FDI", "Kỹ sư phần mềm", 
                   "2018", "2023", "KCN Yên Phong, Bắc Ninh", ""]
    },
    "Hoc_Tap_Cong_Tac_NN": {
        "name": "Du học - Công tác nước ngoài",
        "headers": ["CCCD (*)", "Quốc gia (*)", "Tên trường/Tổ chức (*)", 
                   "Hình thức (Du học/Công tác/Thuê lao động)", "Chuyên ngành/Công việc",
                   "Năm đi", "Năm về", "Nguồn tài trợ", "Ghi chú"],
        "sample": ["001234567890", "Nhật Bản", "Đại học Tokyo", 
                   "Du học", "Thạc sĩ CNTT",
                   "2015", "2019", "Học bổng MEXT", ""]
    },
    "Vi_Pham_NN": {
        "name": "Vi phạm pháp luật ở nước ngoài",
        "headers": ["CCCD (*)", "Quốc gia xảy ra vi phạm (*)", "Năm vi phạm (*)", 
                   "Loại vi phạm", "Nội dung chi tiết", 
                   "Hình thức xử lý", "Tình trạng hiện tại", "Ghi chú"],
        "sample": ["001234567890", "Đài Loan", "2014", 
                   "Cư trú bất hợp pháp", "Ở quá hạn visa 30 ngày", 
                   "Trục xuất", "Đã về nước", ""]
    },
    "Xac_Minh": {
        "name": "Đã từng được xác minh",
        "headers": ["CCCD (*)", "Ngày đề nghị xác minh (*)", "Cơ quan đề nghị (*)", 
                   "Nội dung xác minh", "Cơ quan thực hiện", 
                   "Ngày hoàn thành", "Kết quả", "Ghi chú"],
        "sample": ["001234567890", "15/01/2024", "Sở Công thương", 
                   "Xác minh lý lịch để bổ nhiệm", "PA01", 
                   "30/01/2024", "Đủ điều kiện", ""]
    }
}

def create_excel_template(loai_csxh=None):
    """
    Tạo file Excel mẫu với 5 sheets cho import dữ liệu
    Args:
        loai_csxh: Loại hình CSXH cụ thể (None = tất cả loại)
    Returns: bytes của file Excel
    """
    wb = Workbook()
    
    # Style chung
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header(ws, headers, row=1):
        """Áp dụng style cho header"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 5)
    
    # ========== SHEET 1: ĐỐI TƯỢNG ==========
    ws1 = wb.active
    ws1.title = "1. Đối tượng"
    headers_1 = [
        "CCCD (*)", "Họ và tên (*)", "Ngày sinh (dd/mm/yyyy)", 
        "Giới tính", "Tỉnh/TP", "Xã/Phường",
        "Phân loại nghề nghiệp", "Chi tiết nơi làm việc", "Ghi chú chung"
    ]
    style_header(ws1, headers_1)
    
    # Dòng mẫu
    ws1.append(["001234567890", "Nguyễn Văn A", "01/01/1990", "Nam", 
                "Phú Thọ", "Phường Thanh Miếu", "Cơ quan nhà nước", 
                "Công an tỉnh Phú Thọ", "Ghi chú mẫu"])
    ws1.append(["001234567891", "Trần Thị B", "15/05/1985", "Nữ", 
                "Phú Thọ", "Phường Gia Cẩm", "Tự do", 
                "Buôn bán tự do", ""])
    
    # Ghi chú
    ws1.cell(row=5, column=1, value="Lưu ý: (*) là trường bắt buộc. CCCD phải đủ 12 số.")
    
    # ========== SHEET 2: LIÊN HỆ ==========
    ws2 = wb.create_sheet("2. Liên hệ")
    headers_2 = [
        "CCCD (*)", "Loại liên hệ", "Giá trị (*)", "Ghi chú"
    ]
    style_header(ws2, headers_2)
    ws2.append(["001234567890", "Số điện thoại", "0912345678", "SĐT chính"])
    ws2.append(["001234567890", "Facebook", "facebook.com/nguyenvana", ""])
    ws2.append(["001234567891", "Số điện thoại", "0987654321", ""])
    ws2.cell(row=6, column=1, value="Loại liên hệ: Số điện thoại, Email, Facebook, Zalo, Telegram, Khác")
    
    # ========== SHEET 3: TÀI CHÍNH ==========
    ws3 = wb.create_sheet("3. Tài chính")
    headers_3 = [
        "CCCD (*)", "Ngân hàng", "Số tài khoản (*)", "Chủ tài khoản", "Ghi chú"
    ]
    style_header(ws3, headers_3)
    ws3.append(["001234567890", "Vietcombank", "1234567890123", "NGUYEN VAN A", "TK chính"])
    ws3.append(["001234567890", "Techcombank", "9876543210", "NGUYEN VAN A", "TK phụ"])
    
    # ========== SHEET 4: PHƯƠNG TIỆN ==========
    ws4 = wb.create_sheet("4. Phương tiện")
    headers_4 = [
        "CCCD (*)", "Loại xe", "Biển kiểm soát (*)", "Tên phương tiện", "Ghi chú"
    ]
    style_header(ws4, headers_4)
    ws4.append(["001234567890", "Ô tô", "19A-12345", "Toyota Vios 2022", "Xe cá nhân"])
    ws4.append(["001234567891", "Xe máy", "19B1-67890", "Honda Wave", ""])
    ws4.cell(row=5, column=1, value="Loại xe: Ô tô, Xe máy, Xe tải, Xe khách, Khác")
    
    # ========== SHEET 5: HỒ SƠ CSXH ==========
    if loai_csxh and loai_csxh in CSXH_TEMPLATES:
        # Tạo sheet theo loại cụ thể
        template = CSXH_TEMPLATES[loai_csxh]
        ws5 = wb.create_sheet(f"5. {template['name']}")
        style_header(ws5, template['headers'])
        ws5.append(template['sample'])
        ws5.cell(row=4, column=1, value=f"Loại hình: {loai_csxh}")
    else:
        # Tạo sheet tổng hợp (tất cả loại)
        ws5 = wb.create_sheet("5. Hồ sơ CSXH (Tổng hợp)")
        headers_5 = [
            "CCCD (*)", "Loại hình (*)", 
            "Quốc tịch/Quốc gia", "Tên tổ chức/Người nước ngoài",
            "Thời gian (từ năm)", "Thời gian (đến năm)",
            "Nội dung chi tiết", "Cơ quan xác minh", "Kết quả", "Ghi chú"
        ]
        style_header(ws5, headers_5)
        
        # Mẫu cho từng loại
        ws5.append(["001234567890", "Hon_Nhan_NN", "Trung Quốc", "WANG Xiaoming", 
                    "2020", "", "Kết hôn với công dân TQ", "", "", ""])
        
        ws5.cell(row=4, column=1, value="--- HƯỚNG DẪN LOẠI HÌNH ---")
        ws5.cell(row=5, column=1, value="Hon_Nhan_NN | Lam_Viec_NN | Hoc_Tap_Cong_Tac_NN | Vi_Pham_NN | Xac_Minh")
    
    # Lưu vào buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


# ============================================
# VALIDATE DỮ LIỆU
# ============================================

def validate_excel_data(excel_file):
    """
    Đọc và validate dữ liệu từ file Excel
    Returns: dict với kết quả validate cho từng sheet
    """
    results = {
        'doi_tuong': {'data': None, 'errors': [], 'valid_count': 0, 'error_rows': []},
        'lien_he': {'data': None, 'errors': [], 'valid_count': 0, 'error_rows': []},
        'tai_chinh': {'data': None, 'errors': [], 'valid_count': 0, 'error_rows': []},
        'phuong_tien': {'data': None, 'errors': [], 'valid_count': 0, 'error_rows': []},
        'ho_so_dac_thu': {'data': None, 'errors': [], 'valid_count': 0, 'error_rows': []},
    }
    
    try:
        # Đọc tất cả sheets
        xls = pd.ExcelFile(excel_file)
        sheet_names = xls.sheet_names
        
        # Lấy danh sách CCCD đã tồn tại trong DB
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT cccd FROM doi_tuong")
        existing_cccds = set(row[0] for row in cursor.fetchall())
        conn.close()
        
        # ===== SHEET 1: ĐỐI TƯỢNG =====
        if len(sheet_names) >= 1:
            df = pd.read_excel(xls, sheet_name=0, skiprows=0)
            df = df.iloc[1:]  # Bỏ dòng mẫu
            df = df.dropna(how='all')  # Bỏ dòng trống
            
            if len(df) > 0:
                # Chuẩn hóa tên cột
                df.columns = ['cccd', 'ho_ten', 'ngay_sinh', 'gioi_tinh', 
                             'dia_chi_tinh', 'dia_chi_xa', 'phan_loai_nghe_nghiep',
                             'chi_tiet_nghe_nghiep', 'ghi_chu_chung']
                
                errors = []
                valid_rows = []
                new_cccds = set()
                
                for idx, row in df.iterrows():
                    row_errors = []
                    cccd = str(row['cccd']).strip() if pd.notna(row['cccd']) else ""
                    
                    # Validate CCCD (phải đủ 12 số)
                    if not cccd:
                        row_errors.append(f"Dòng {idx+1}: Thiếu CCCD")
                    elif not cccd.isdigit():
                        row_errors.append(f"Dòng {idx+1}: CCCD chỉ được chứa số")
                    elif len(cccd) != 12:
                        row_errors.append(f"Dòng {idx+1}: CCCD phải đủ 12 số (hiện có {len(cccd)} số)")
                    elif cccd in existing_cccds:
                        row_errors.append(f"Dòng {idx+1}: CCCD {cccd} đã tồn tại trong hệ thống")
                    elif cccd in new_cccds:
                        row_errors.append(f"Dòng {idx+1}: CCCD {cccd} bị trùng trong file")
                    
                    # Validate Họ tên (bắt buộc)
                    ho_ten = str(row['ho_ten']).strip() if pd.notna(row['ho_ten']) else ""
                    if not ho_ten:
                        row_errors.append(f"Dòng {idx+1}: Thiếu Họ và tên")
                    
                    # Validate Ngày sinh (định dạng dd/mm/yyyy hoặc datetime)
                    ngay_sinh = row['ngay_sinh']
                    if pd.notna(ngay_sinh):
                        try:
                            if isinstance(ngay_sinh, str):
                                # Kiểm tra định dạng dd/mm/yyyy
                                if not re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', ngay_sinh.strip()):
                                    row_errors.append(f"Dòng {idx+1}: Ngày sinh sai định dạng (cần dd/mm/yyyy)")
                                else:
                                    datetime.strptime(ngay_sinh.strip(), '%d/%m/%Y')
                        except ValueError:
                            row_errors.append(f"Dòng {idx+1}: Ngày sinh không hợp lệ")
                    
                    # Validate Giới tính
                    gioi_tinh = str(row['gioi_tinh']).strip() if pd.notna(row['gioi_tinh']) else ""
                    if gioi_tinh and gioi_tinh not in GIOI_TINH_OPTIONS:
                        row_errors.append(f"Dòng {idx+1}: Giới tính '{gioi_tinh}' không hợp lệ (chỉ: {', '.join(GIOI_TINH_OPTIONS)})")
                    
                    # Validate Tỉnh/TP
                    dia_chi_tinh = str(row['dia_chi_tinh']).strip() if pd.notna(row['dia_chi_tinh']) else ""
                    if dia_chi_tinh and dia_chi_tinh not in TINH_OPTIONS:
                        row_errors.append(f"Dòng {idx+1}: Tỉnh/TP '{dia_chi_tinh}' không hợp lệ (chỉ: {', '.join(TINH_OPTIONS)})")
                    
                    # Validate Xã/Phường (phải nằm trong danh sách 105 nếu tỉnh là Phú Thọ)
                    dia_chi_xa = str(row['dia_chi_xa']).strip() if pd.notna(row['dia_chi_xa']) else ""
                    if dia_chi_tinh == "Phú Thọ" and dia_chi_xa:
                        if dia_chi_xa not in DANH_SACH_XA_PHU_THO:
                            row_errors.append(f"Dòng {idx+1}: Xã/Phường '{dia_chi_xa}' không nằm trong danh sách 105 xã/phường Phú Thọ")
                    
                    # Validate Phân loại nghề nghiệp
                    phan_loai = str(row['phan_loai_nghe_nghiep']).strip() if pd.notna(row['phan_loai_nghe_nghiep']) else ""
                    if phan_loai and phan_loai not in PHAN_LOAI_NGHE_NGHIEP_OPTIONS:
                        row_errors.append(f"Dòng {idx+1}: Phân loại nghề nghiệp '{phan_loai}' không hợp lệ")
                    
                    if row_errors:
                        errors.extend(row_errors)
                        # Lưu dòng lỗi với lý do
                        error_row = row.to_dict()
                        error_row['LY_DO_LOI'] = '; '.join([e.split(': ', 1)[1] if ': ' in e else e for e in row_errors])
                        results['doi_tuong']['error_rows'].append(error_row)
                    else:
                        new_cccds.add(cccd)
                        valid_rows.append(row)
                
                results['doi_tuong']['data'] = pd.DataFrame(valid_rows) if valid_rows else None
                results['doi_tuong']['errors'] = errors
                results['doi_tuong']['valid_count'] = len(valid_rows)
                results['doi_tuong']['new_cccds'] = new_cccds
        
        # ===== SHEET 2: LIÊN HỆ =====
        valid_cccds = existing_cccds.union(results['doi_tuong'].get('new_cccds', set()))
        
        if len(sheet_names) >= 2:
            df = pd.read_excel(xls, sheet_name=1, skiprows=0)
            df = df.iloc[1:]
            df = df.dropna(how='all')
            
            if len(df) > 0:
                df.columns = ['cccd', 'loai_lien_he', 'gia_tri', 'ghi_chu']
                errors = []
                valid_rows = []
                
                for idx, row in df.iterrows():
                    row_errors = []
                    cccd = str(row['cccd']).strip() if pd.notna(row['cccd']) else ""
                    loai_lien_he = str(row['loai_lien_he']).strip() if pd.notna(row['loai_lien_he']) else ""
                    gia_tri = str(row['gia_tri']).strip() if pd.notna(row['gia_tri']) else ""
                    
                    if cccd not in valid_cccds:
                        row_errors.append(f"Dòng {idx+1}: CCCD {cccd} không tồn tại")
                    if not gia_tri:
                        row_errors.append(f"Dòng {idx+1}: Thiếu giá trị liên hệ")
                    if loai_lien_he and loai_lien_he not in LOAI_LIEN_HE_OPTIONS + ["Khác", "Số điện thoại", "Facebook", "Zalo", "Telegram", "Email"]:
                        row_errors.append(f"Dòng {idx+1}: Loại liên hệ '{loai_lien_he}' không hợp lệ")
                    
                    if row_errors:
                        errors.extend(row_errors)
                        error_row = row.to_dict()
                        error_row['LY_DO_LOI'] = '; '.join([e.split(': ', 1)[1] if ': ' in e else e for e in row_errors])
                        results['lien_he']['error_rows'].append(error_row)
                    else:
                        valid_rows.append(row)
                
                results['lien_he']['data'] = pd.DataFrame(valid_rows) if valid_rows else None
                results['lien_he']['errors'] = errors
                results['lien_he']['valid_count'] = len(valid_rows)
        
        # ===== SHEET 3: TÀI CHÍNH =====
        if len(sheet_names) >= 3:
            df = pd.read_excel(xls, sheet_name=2, skiprows=0)
            df = df.iloc[1:]
            df = df.dropna(how='all')
            
            if len(df) > 0:
                df.columns = ['cccd', 'ngan_hang', 'so_tai_khoan', 'chu_tai_khoan', 'ghi_chu']
                errors = []
                valid_rows = []
                
                for idx, row in df.iterrows():
                    row_errors = []
                    cccd = str(row['cccd']).strip() if pd.notna(row['cccd']) else ""
                    ngan_hang = str(row['ngan_hang']).strip() if pd.notna(row['ngan_hang']) else ""
                    so_tai_khoan = str(row['so_tai_khoan']).strip() if pd.notna(row['so_tai_khoan']) else ""
                    
                    if cccd not in valid_cccds:
                        row_errors.append(f"Dòng {idx+1}: CCCD {cccd} không tồn tại")
                    if not so_tai_khoan:
                        row_errors.append(f"Dòng {idx+1}: Thiếu số tài khoản")
                    if ngan_hang and ngan_hang not in DANH_SACH_NGAN_HANG:
                        row_errors.append(f"Dòng {idx+1}: Ngân hàng '{ngan_hang}' không nằm trong danh sách chuẩn")
                    
                    if row_errors:
                        errors.extend(row_errors)
                        error_row = row.to_dict()
                        error_row['LY_DO_LOI'] = '; '.join([e.split(': ', 1)[1] if ': ' in e else e for e in row_errors])
                        results['tai_chinh']['error_rows'].append(error_row)
                    else:
                        valid_rows.append(row)
                
                results['tai_chinh']['data'] = pd.DataFrame(valid_rows) if valid_rows else None
                results['tai_chinh']['errors'] = errors
                results['tai_chinh']['valid_count'] = len(valid_rows)
        
        # ===== SHEET 4: PHƯƠNG TIỆN =====
        if len(sheet_names) >= 4:
            df = pd.read_excel(xls, sheet_name=3, skiprows=0)
            df = df.iloc[1:]
            df = df.dropna(how='all')
            
            if len(df) > 0:
                df.columns = ['cccd', 'loai_xe', 'bien_kiem_soat', 'ten_phuong_tien', 'ghi_chu']
                errors = []
                valid_rows = []
                
                for idx, row in df.iterrows():
                    row_errors = []
                    cccd = str(row['cccd']).strip() if pd.notna(row['cccd']) else ""
                    loai_xe = str(row['loai_xe']).strip() if pd.notna(row['loai_xe']) else ""
                    bien_kiem_soat = str(row['bien_kiem_soat']).strip() if pd.notna(row['bien_kiem_soat']) else ""
                    
                    if cccd not in valid_cccds:
                        row_errors.append(f"Dòng {idx+1}: CCCD {cccd} không tồn tại")
                    if not bien_kiem_soat:
                        row_errors.append(f"Dòng {idx+1}: Thiếu biển kiểm soát")
                    if loai_xe and loai_xe not in LOAI_XE_OPTIONS + ["Xe tải", "Xe khách", "Khác"]:
                        row_errors.append(f"Dòng {idx+1}: Loại xe '{loai_xe}' không hợp lệ")
                    
                    if row_errors:
                        errors.extend(row_errors)
                        error_row = row.to_dict()
                        error_row['LY_DO_LOI'] = '; '.join([e.split(': ', 1)[1] if ': ' in e else e for e in row_errors])
                        results['phuong_tien']['error_rows'].append(error_row)
                    else:
                        valid_rows.append(row)
                
                results['phuong_tien']['data'] = pd.DataFrame(valid_rows) if valid_rows else None
                results['phuong_tien']['errors'] = errors
                results['phuong_tien']['valid_count'] = len(valid_rows)
        
        # ===== SHEET 5: HỒ SƠ CSXH =====
        valid_loai_hinh = ['Hon_Nhan_NN', 'Lam_Viec_NN', 'Hoc_Tap_Cong_Tac_NN', 'Vi_Pham_NN', 'Xac_Minh']
        
        if len(sheet_names) >= 5:
            df = pd.read_excel(xls, sheet_name=4, skiprows=0)
            df = df.iloc[1:]
            df = df.dropna(how='all')
            
            # Lọc bỏ các dòng hướng dẫn (bắt đầu bằng "---" hoặc "-")
            if len(df) > 0:
                df = df[~df.iloc[:, 0].astype(str).str.startswith('---')]
                df = df[~df.iloc[:, 0].astype(str).str.startswith('-')]
            
            if len(df) > 0:
                df.columns = ['cccd', 'loai_hinh', 'quoc_tich', 'ten_to_chuc',
                             'thoi_gian_tu', 'thoi_gian_den', 'noi_dung_chi_tiet',
                             'co_quan_xm', 'ket_qua', 'ghi_chu']
                errors = []
                valid_rows = []
                
                for idx, row in df.iterrows():
                    row_errors = []
                    cccd = str(row['cccd']).strip() if pd.notna(row['cccd']) else ""
                    loai_hinh = str(row['loai_hinh']).strip() if pd.notna(row['loai_hinh']) else ""
                    quoc_tich = str(row['quoc_tich']).strip() if pd.notna(row['quoc_tich']) else ""
                    
                    if cccd not in valid_cccds:
                        row_errors.append(f"Dòng {idx+1}: CCCD {cccd} không tồn tại")
                    if not loai_hinh:
                        row_errors.append(f"Dòng {idx+1}: Thiếu loại hình")
                    elif loai_hinh not in valid_loai_hinh:
                        row_errors.append(f"Dòng {idx+1}: Loại hình '{loai_hinh}' không hợp lệ (cần: {', '.join(valid_loai_hinh)})")
                    
                    # Kiểm tra quốc gia (nếu là loại liên quan đến nước ngoài)
                    if loai_hinh in ['Hon_Nhan_NN', 'Lam_Viec_NN', 'Hoc_Tap_Cong_Tac_NN', 'Vi_Pham_NN']:
                        if quoc_tich and quoc_tich not in DANH_SACH_QUOC_GIA:
                            row_errors.append(f"Dòng {idx+1}: Quốc gia '{quoc_tich}' không nằm trong danh sách chuẩn")
                    
                    if row_errors:
                        errors.extend(row_errors)
                        error_row = row.to_dict()
                        error_row['LY_DO_LOI'] = '; '.join([e.split(': ', 1)[1] if ': ' in e else e for e in row_errors])
                        results['ho_so_dac_thu']['error_rows'].append(error_row)
                    else:
                        valid_rows.append(row)
                
                results['ho_so_dac_thu']['data'] = pd.DataFrame(valid_rows) if valid_rows else None
                results['ho_so_dac_thu']['errors'] = errors
                results['ho_so_dac_thu']['valid_count'] = len(valid_rows)
        
    except Exception as e:
        results['doi_tuong']['errors'].append(f"Lỗi đọc file: {str(e)}")
    
    return results


# ============================================
# XUẤT FILE CÁC DÒNG LỖI
# ============================================

def export_error_excel(validation_results):
    """
    Tạo file Excel chứa các dòng lỗi với cột lý do để user fix
    Returns: bytes của file Excel hoặc None nếu không có lỗi
    """
    wb = Workbook()
    
    # Style cho header
    header_font = Font(bold=True, color="FFFFFF")
    error_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    
    has_errors = False
    first_sheet = True
    
    # Sheet tên và dữ liệu
    sheet_configs = [
        ('1. Đối tượng - LỖI', 'doi_tuong'),
        ('2. Liên hệ - LỖI', 'lien_he'),
        ('3. Tài chính - LỖI', 'tai_chinh'),
        ('4. Phương tiện - LỖI', 'phuong_tien'),
        ('5. CSXH - LỖI', 'ho_so_dac_thu'),
    ]
    
    for sheet_name, key in sheet_configs:
        error_rows = validation_results.get(key, {}).get('error_rows', [])
        if error_rows:
            has_errors = True
            
            if first_sheet:
                ws = wb.active
                ws.title = sheet_name
                first_sheet = False
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Tạo DataFrame từ error_rows
            df = pd.DataFrame(error_rows)
            
            # Đảm bảo cột LY_DO_LOI ở cuối
            if 'LY_DO_LOI' in df.columns:
                cols = [c for c in df.columns if c != 'LY_DO_LOI'] + ['LY_DO_LOI']
                df = df[cols]
            
            # Viết header
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                if col_name == 'LY_DO_LOI':
                    cell.fill = error_fill
                else:
                    cell.fill = header_fill
                cell.font = header_font
                ws.column_dimensions[cell.column_letter].width = max(15, len(str(col_name)) + 5)
            
            # Viết dữ liệu
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row.values, 1):
                    cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
                    # Highlight cột lý do lỗi
                    if df.columns[col_idx - 1] == 'LY_DO_LOI':
                        cell.font = Font(color="dc3545", bold=True)
    
    if not has_errors:
        return None
    
    # Lưu vào buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


# ============================================
# BULK INSERT VỚI TRANSACTION
# ============================================

def bulk_import_all(validated_data):
    """
    Import tất cả dữ liệu đã validate vào database với transaction
    Sử dụng executemany để tối ưu hiệu năng
    Nếu có lỗi bất kỳ đâu -> rollback toàn bộ
    Returns: (success, message, stats)
    """
    conn = get_connection()
    cursor = conn.cursor()
    
    stats = {
        'doi_tuong': 0,
        'lien_he': 0,
        'tai_chinh': 0,
        'phuong_tien': 0,
        'ho_so_dac_thu': 0
    }

    # Helper function to get value safely
    def get_val(row, col):
        val = row.get(col)
        return str(val).strip() if pd.notna(val) else None
    
    try:
        # Bắt đầu transaction
        conn.execute("BEGIN TRANSACTION")
        
        # ===== INSERT ĐỐI TƯỢNG =====
        if validated_data['doi_tuong']['data'] is not None:
            df = validated_data['doi_tuong']['data']
            data_list = []

            for _, row in df.iterrows():
                # Xử lý ngày sinh
                ngay_sinh = None
                raw_ns = row.get('ngay_sinh')
                if pd.notna(raw_ns):
                    try:
                        if isinstance(raw_ns, str):
                            # Validate đã đảm bảo format hoặc data sạch, nhưng vẫn cần try/except để an toàn
                            ngay_sinh = datetime.strptime(raw_ns, '%d/%m/%Y').strftime('%Y-%m-%d')
                        elif hasattr(raw_ns, 'strftime'):
                            ngay_sinh = raw_ns.strftime('%Y-%m-%d')
                    except ValueError as e:
                        # Không nuốt lỗi (silent fail)
                        raise ValueError(f"Lỗi xử lý ngày sinh cho CCCD {row.get('cccd')}: {str(e)}")

                data_list.append((
                    str(row['cccd']).strip(),
                    get_val(row, 'ho_ten'),
                    ngay_sinh,
                    get_val(row, 'gioi_tinh'),
                    get_val(row, 'dia_chi_tinh') or 'Phú Thọ',
                    get_val(row, 'dia_chi_xa'),
                    get_val(row, 'phan_loai_nghe_nghiep'),
                    get_val(row, 'chi_tiet_nghe_nghiep'),
                    get_val(row, 'ghi_chu_chung')
                ))

            if data_list:
                cursor.executemany("""
                    INSERT INTO doi_tuong 
                    (cccd, ho_ten, ngay_sinh, gioi_tinh, dia_chi_tinh, dia_chi_xa, 
                     phan_loai_nghe_nghiep, chi_tiet_nghe_nghiep, ghi_chu_chung)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, data_list)
                stats['doi_tuong'] = len(data_list)
        
        # ===== INSERT LIÊN HỆ =====
        if validated_data['lien_he']['data'] is not None:
            df = validated_data['lien_he']['data']
            data_list = []
            for _, row in df.iterrows():
                data_list.append((
                    str(row['cccd']).strip(),
                    get_val(row, 'loai_lien_he'),
                    str(row['gia_tri']).strip(),
                    get_val(row, 'ghi_chu')
                ))

            if data_list:
                cursor.executemany("""
                    INSERT INTO lien_he (cccd, loai_lien_he, gia_tri, ghi_chu)
                    VALUES (?, ?, ?, ?)
                """, data_list)
                stats['lien_he'] = len(data_list)
        
        # ===== INSERT TÀI CHÍNH =====
        if validated_data['tai_chinh']['data'] is not None:
            df = validated_data['tai_chinh']['data']
            data_list = []
            for _, row in df.iterrows():
                data_list.append((
                    str(row['cccd']).strip(),
                    get_val(row, 'ngan_hang'),
                    str(row['so_tai_khoan']).strip(),
                    get_val(row, 'chu_tai_khoan'),
                    get_val(row, 'ghi_chu')
                ))

            if data_list:
                cursor.executemany("""
                    INSERT INTO tai_chinh (cccd, ngan_hang, so_tai_khoan, chu_tai_khoan, ghi_chu)
                    VALUES (?, ?, ?, ?, ?)
                """, data_list)
                stats['tai_chinh'] = len(data_list)
        
        # ===== INSERT PHƯƠNG TIỆN =====
        if validated_data['phuong_tien']['data'] is not None:
            df = validated_data['phuong_tien']['data']
            data_list = []
            for _, row in df.iterrows():
                data_list.append((
                    str(row['cccd']).strip(),
                    get_val(row, 'loai_xe'),
                    str(row['bien_kiem_soat']).strip(),
                    get_val(row, 'ten_phuong_tien'),
                    get_val(row, 'ghi_chu')
                ))

            if data_list:
                cursor.executemany("""
                    INSERT INTO phuong_tien (cccd, loai_xe, bien_kiem_soat, ten_phuong_tien, ghi_chu)
                    VALUES (?, ?, ?, ?, ?)
                """, data_list)
                stats['phuong_tien'] = len(data_list)
        
        # ===== INSERT HỒ SƠ CSXH =====
        if validated_data['ho_so_dac_thu']['data'] is not None:
            df = validated_data['ho_so_dac_thu']['data']
            data_list = []
            import json
            for _, row in df.iterrows():
                noi_dung_dict = {
                    'quoc_tich': get_val(row, 'quoc_tich') or '',
                    'ten_to_chuc': get_val(row, 'ten_to_chuc') or '',
                    'thoi_gian_tu': get_val(row, 'thoi_gian_tu') or '',
                    'thoi_gian_den': get_val(row, 'thoi_gian_den') or '',
                    'noi_dung': get_val(row, 'noi_dung_chi_tiet') or '',
                    'co_quan_xm': get_val(row, 'co_quan_xm') or '',
                    'ket_qua': get_val(row, 'ket_qua') or '',
                }
                
                data_list.append((
                    str(row['cccd']).strip(),
                    str(row['loai_hinh']).strip(),
                    json.dumps(noi_dung_dict, ensure_ascii=False),
                    get_val(row, 'ghi_chu')
                ))

            if data_list:
                cursor.executemany("""
                    INSERT INTO ho_so_dac_thu (cccd, loai_hinh, noi_dung_chi_tiet, ghi_chu)
                    VALUES (?, ?, ?, ?)
                """, data_list)
                stats['ho_so_dac_thu'] = len(data_list)
        
        # Commit transaction
        conn.commit()
        conn.close()
        
        total = sum(stats.values())
        return True, f"Import thành công {total} bản ghi!", stats
        
    except Exception as e:
        # Rollback nếu có lỗi
        conn.rollback()
        conn.close()
        return False, f"Lỗi import tại bước SQL: {str(e)}", stats
